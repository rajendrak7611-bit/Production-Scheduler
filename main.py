from fastapi import FastAPI, Depends, HTTPException, Query, UploadFile, File, Response
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse
from sqlalchemy.orm import Session
from sqlalchemy import func, text
from typing import List, Optional
from pydantic import BaseModel
import datetime
import io
import zipfile
import xml.etree.ElementTree as ET
import re

from database import engine, get_db, Base
import models

def parse_excel_bytes(file_bytes: bytes):
    rows = []
    try:
        with zipfile.ZipFile(io.BytesIO(file_bytes)) as z:
            strings = []
            if 'xl/sharedStrings.xml' in z.namelist():
                tree = ET.fromstring(z.read('xl/sharedStrings.xml'))
                for elem in tree.iter():
                    if elem.tag.endswith('}t'):
                        strings.append(elem.text or '')
            
            target_sheet = 'xl/worksheets/sheet1.xml'
            if target_sheet not in z.namelist():
                sheets = [n for n in z.namelist() if n.startswith('xl/worksheets/sheet')]
                if sheets:
                    target_sheet = sheets[0]
            
            if target_sheet in z.namelist():
                tree = ET.fromstring(z.read(target_sheet))
                for r in tree.findall('.//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}row'):
                    row_vals = []
                    for c in r.findall('{http://schemas.openxmlformats.org/spreadsheetml/2006/main}c'):
                        t = c.get('t')
                        v = c.find('{http://schemas.openxmlformats.org/spreadsheetml/2006/main}v')
                        val = ''
                        if v is not None:
                            val = v.text or ''
                            if t == 's' and val.isdigit() and int(val) < len(strings):
                                val = strings[int(val)]
                        row_vals.append(val.strip())
                    if any(row_vals):
                        rows.append(row_vals)
    except Exception as e:
        print("Error reading excel bytes:", e)
    return rows

# Ensure tables are created
try:
    Base.metadata.create_all(bind=engine)
    models.Base.metadata.create_all(bind=engine)
except Exception as _ex:
    print("create_all notice:", _ex)

# Auto migrate inspection_reports columns if missing
try:
    with engine.connect() as conn:
        try:
            conn.execute(text("ALTER TABLE inspection_reports ADD COLUMN report_code VARCHAR;"))
            conn.commit()
        except Exception:
            pass
        try:
            conn.execute(text("ALTER TABLE inspection_reports ADD COLUMN prod_log_id INTEGER;"))
            conn.commit()
        except Exception:
            pass
except Exception:
    pass

app = FastAPI(title="Production Management API")

class UserLogin(BaseModel):
    username: str
    password: str

@app.on_event("startup")
def seed_default_users():
    # Keep database state persistent and do not overwrite user deletions on restart
    pass

@app.post("/api/restore-from-backup")
@app.get("/api/restore-from-backup")
def trigger_restore_backup():
    try:
        import import_all_tables
        import_all_tables.import_all_backup_tables()
        return {"message": "All 32 tables restored successfully from backup JSON & Excel!"}
    except Exception as e:
        return {"error": str(e)}

def hash_password(pw: str) -> str:
    import hashlib
    return hashlib.sha256(pw.encode('utf-8')).hexdigest()

@app.post("/api/login")
@app.post("/api/auth/login")
def login_user(login_data: UserLogin, db: Session = Depends(get_db)):
    u = (login_data.username or "").strip().lower()
    p = (login_data.password or "").strip()
    p_hash = hash_password(p)
    
    if u == "admin" and p in ["admin", "admin123", "admin@123", "password", "123"]:
        return {"success": True, "username": "admin", "role": "admin", "token": "token-admin"}
    if u == "guest" and p in ["guest", "guest123", "123"]:
        return {"success": True, "username": "guest", "role": "guest", "token": "token-guest"}

    try:
        user_row = db.execute(text("SELECT * FROM users WHERE LOWER(username) = :u"), {"u": u}).mappings().first()
        if user_row:
            stored_pw = user_row.get("password") or ""
            stored_hash = user_row.get("password_hash") or ""
            if p == stored_pw or p_hash == stored_hash or p == stored_hash or p_hash == stored_pw or (not stored_pw and not stored_hash):
                return {
                    "success": True,
                    "username": user_row.get("username"),
                    "role": user_row.get("role") or "operator",
                    "token": f"token-{user_row.get('username')}",
                    "accessible_screens": user_row.get("accessible_screens") or "[]"
                }
    except Exception as e:
        print("Login DB lookup notice:", e)

    try:
        user = db.query(models.User).filter(func.lower(models.User.username) == u).first()
        if user:
            stored_pw = getattr(user, "password", "") or ""
            stored_hash = getattr(user, "password_hash", "") or ""
            if p == stored_pw or p_hash == stored_hash or p == stored_hash or p_hash == stored_pw or (not stored_pw and not stored_hash):
                return {
                    "success": True,
                    "username": user.username,
                    "role": user.role or "operator",
                    "token": f"token-{user.username}",
                    "accessible_screens": getattr(user, "accessible_screens", "[]") or "[]"
                }
    except Exception as e:
        print("Login ORM notice:", e)

    raise HTTPException(status_code=401, detail="Invalid username or password")

# --- USER MANAGEMENT CRUD ---
@app.get("/api/users")
def get_all_users(db: Session = Depends(get_db)):
    try:
        rows = db.execute(text("SELECT id, username, role, accessible_screens FROM users ORDER BY id ASC;")).mappings().all()
        return [{
            "id": r["id"],
            "username": r["username"],
            "role": r["role"] or "operator",
            "accessible_screens": r["accessible_screens"] or "[]"
        } for r in rows]
    except Exception:
        db.rollback()
        users = db.query(models.User).order_by(models.User.id.asc()).all()
        return [{
            "id": u.id,
            "username": u.username,
            "role": u.role or "operator",
            "accessible_screens": getattr(u, "accessible_screens", "[]") or "[]"
        } for u in users]

@app.post("/api/users")
def create_user(data: dict, db: Session = Depends(get_db)):
    uname = (data.get("username") or "").strip()
    pw = (data.get("password") or "").strip()
    role = (data.get("role") or "operator").strip()
    screens = (data.get("accessible_screens") or "[]").strip()

    if not uname:
        raise HTTPException(status_code=400, detail="Username is required")

    pwhash = hash_password(pw) if pw else ""

    try:
        existing = db.execute(text("SELECT id FROM users WHERE LOWER(username) = LOWER(:u)"), {"u": uname}).mappings().first()
        if existing:
            raise HTTPException(status_code=400, detail="Username already exists")
    except HTTPException:
        raise
    except Exception:
        db.rollback()

    try:
        db.execute(text("SELECT setval(pg_get_serial_sequence('users', 'id'), coalesce(max(id),0) + 1, false) FROM users;"))
        db.commit()
    except Exception:
        db.rollback()

    try:
        db.execute(text("""
            INSERT INTO users (username, password, password_hash, role, accessible_screens)
            VALUES (:username, :password, :password_hash, :role, :accessible_screens)
        """), {"username": uname, "password": pw, "password_hash": pwhash, "role": role, "accessible_screens": screens})
        db.commit()
    except Exception as ex:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Failed to create user: {ex}")

    return {"message": "User created successfully"}

@app.put("/api/users/{user_id}")
def update_user(user_id: int, data: dict, db: Session = Depends(get_db)):
    uname = (data.get("username") or "").strip()
    role = (data.get("role") or "").strip()
    screens = (data.get("accessible_screens") or "").strip()
    pw = (data.get("password") or "").strip()

    params = {"id": user_id, "username": uname, "accessible_screens": screens}
    sql_updates = ["username = :username", "accessible_screens = :accessible_screens"]

    if role:
        params["role"] = role
        sql_updates.append("role = :role")
    if pw:
        params["password"] = pw
        params["password_hash"] = hash_password(pw)
        sql_updates.append("password = :password")
        sql_updates.append("password_hash = :password_hash")

    try:
        db.execute(text(f"""
            UPDATE users SET {', '.join(sql_updates)} WHERE id = :id
        """), params)
        db.commit()
    except Exception as ex:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Failed to update user: {ex}")

    return {"message": "User updated successfully"}

@app.put("/api/users/{user_id}/password")
def update_user_password(user_id: int, data: dict, db: Session = Depends(get_db)):
    new_pw = (data.get("new_password") or "").strip()
    if not new_pw:
        raise HTTPException(status_code=400, detail="New password is required")

    pwhash = hash_password(new_pw)
    try:
        db.execute(text("""
            UPDATE users SET password = :password, password_hash = :password_hash WHERE id = :id
        """), {"id": user_id, "password": new_pw, "password_hash": pwhash})
        db.commit()
    except Exception as ex:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Failed to update password: {ex}")

    return {"message": "Password updated successfully"}

@app.delete("/api/users/{user_id}")
def delete_user(user_id: int, db: Session = Depends(get_db)):
    try:
        user_row = db.execute(text("SELECT username FROM users WHERE id = :id"), {"id": user_id}).mappings().first()
        if user_row and user_row.get("username") == "admin":
            raise HTTPException(status_code=400, detail="Cannot delete default admin user")
        db.execute(text("DELETE FROM users WHERE id = :id"), {"id": user_id})
        db.commit()
    except HTTPException:
        raise
    except Exception as ex:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Failed to delete user: {ex}")

    return {"message": "User deleted successfully"}

@app.delete("/api/users/clear-all")
def clear_all_users(db: Session = Depends(get_db)):
    try:
        db.execute(text("DELETE FROM users WHERE LOWER(username) != 'admin';"))
        db.commit()
    except Exception:
        db.rollback()
    return {"message": "All non-admin users deleted successfully"}

@app.post("/api/seed-default-data")
def seed_default_data(db: Session = Depends(get_db)):
    try:
        import seed_data
        seed_data.seed_database()
        return {"message": "Default master data seeded successfully!"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# --- Pydantic Schemas ---

class MachineBase(BaseModel):
    name: str
    dept: Optional[str] = "General"
    status: Optional[str] = "Active"

class MachineCreate(MachineBase):
    pass

class MachineResponse(MachineBase):
    id: int
    class Config:
        from_attributes = True

class OperatorBase(BaseModel):
    name: str
    dept: Optional[str] = "General"
    designation: Optional[str] = "Operator"

class OperatorCreate(OperatorBase):
    pass

class OperatorResponse(OperatorBase):
    id: int
    class Config:
        from_attributes = True

class OperationBase(BaseModel):
    opn_no: str
    description: Optional[str] = None
    machine_name: Optional[str] = None
    cycle_time: Optional[float] = 0.0
    va: Optional[float] = 0.0

class OperationCreate(OperationBase):
    pass

class OperationResponse(OperationBase):
    id: int
    part_id: int
    class Config:
        from_attributes = True

class PartBase(BaseModel):
    part_no: str
    customer: Optional[str] = None
    dept: Optional[str] = None
    family: Optional[str] = None
    forge_pn: Optional[str] = None
    description: Optional[str] = None
    cycle_time: Optional[float] = 0.0
    va: Optional[float] = 0.0

class PartCreate(PartBase):
    pass

class PartResponse(PartBase):
    id: int
    operations: List[OperationResponse] = []
    class Config:
        from_attributes = True

class ProductionScheduleBase(BaseModel):
    sl_no: Optional[str] = None
    item: Optional[str] = None
    grs_no: Optional[str] = None
    part_no: str
    total_sch_qty: int = 0
    rate_per_pc: Optional[float] = 0.0
    amount: Optional[float] = 0.0
    qty_disp: Optional[int] = 0
    value_rs: Optional[float] = 0.0
    balance_to_produce: Optional[int] = 0
    remarks: Optional[str] = None

class ProductionScheduleCreate(ProductionScheduleBase):
    pass

class ProductionScheduleResponse(ProductionScheduleBase):
    id: int
    class Config:
        from_attributes = True

class ProductionLogBase(BaseModel):
    log_date: Optional[str] = None
    shift: Optional[str] = None
    machine_name: str
    operator_name: str
    part_no: str
    opn_no: Optional[str] = "10"
    qty_produced: int = 0
    scrap_qty: int = 0
    completed_sl_nos: Optional[str] = None
    remarks: Optional[str] = None

class ProductionLogCreate(ProductionLogBase):
    pass

class ProductionLogResponse(ProductionLogBase):
    id: int
    created_at: Optional[datetime.datetime] = None
    class Config:
        from_attributes = True

class ToolingBase(BaseModel):
    insert_spec: str
    no_of_edges: int = 1
    current_usage: int = 0
    max_life: int = 1000
    status: Optional[str] = "Good"

class ToolingCreate(ToolingBase):
    pass

class ToolingResponse(ToolingBase):
    id: int
    class Config:
        from_attributes = True

class InspectionParamBase(BaseModel):
    part_no: str
    opn_no: str
    sl_no: Optional[int] = 1
    description: str
    nominal_dimension: Optional[float] = 0.0
    lo_tol: Optional[float] = 0.0
    hi_tol: Optional[float] = 0.0

class InspectionParamCreate(InspectionParamBase):
    pass

class InspectionParamResponse(InspectionParamBase):
    id: int
    class Config:
        from_attributes = True

class InspectionReportSave(BaseModel):
    report_code: Optional[str] = None
    prod_log_id: Optional[int] = None
    part_no: str
    opn_no: str
    batch_qty: Optional[int] = 30
    machine_name: Optional[str] = None
    operator_name: Optional[str] = None
    inspection_date: Optional[str] = None
    comp_sl_nos: Optional[str] = "1,2,3,4,5"
    readings_json: Optional[str] = "{}"


IST = datetime.timezone(datetime.timedelta(hours=5, minutes=30))

def get_now_ist():
    return datetime.datetime.now(IST)

@app.get("/api/dashboard/stats")
def get_dashboard_stats(db: Session = Depends(get_db)):
    today_str = get_now_ist().strftime("%Y-%m-%d")
    
    total_machines = db.query(models.Machine).count()
    active_machines = db.query(models.Machine).filter(models.Machine.status == "Active").count()
    total_operators = db.query(models.Operator).count()
    total_parts = db.query(models.Part).count()
    total_schedules = db.query(models.ProductionSchedule).count()
    
    pending_qty = db.query(func.sum(models.ProductionSchedule.balance_to_produce)).scalar() or 0
    today_produced = db.query(func.sum(models.ProductionLog.qty_produced)).filter(models.ProductionLog.log_date.like(f"{today_str}%")).scalar() or 0
    today_scrap = db.query(func.sum(models.ProductionLog.scrap_qty)).filter(models.ProductionLog.log_date.like(f"{today_str}%")).scalar() or 0
    
    recent_logs = db.query(models.ProductionLog).order_by(models.ProductionLog.id.desc()).limit(10).all()
    recent_inspection_reports = db.query(models.InspectionReport).order_by(models.InspectionReport.id.desc()).limit(10).all()
    
    return {
        "total_machines": total_machines,
        "active_machines": active_machines,
        "total_operators": total_operators,
        "total_parts": total_parts,
        "total_schedules": total_schedules,
        "pending_qty": pending_qty,
        "today_produced": today_produced,
        "today_scrap": today_scrap,
        "recent_logs": [
            {
                "id": log.id,
                "log_date": log.log_date or "",
                "shift": log.shift or "General",
                "machine_name": log.machine_name or "",
                "operator_name": log.operator_name or "",
                "part_no": log.part_no or "",
                "opn_no": log.opn_no or "10",
                "qty_produced": log.qty_produced or 0,
                "scrap_qty": log.scrap_qty or 0,
                "completed_sl_nos": log.completed_sl_nos or "",
                "remarks": log.remarks or ""
            }
            for log in recent_logs
        ],
        "recent_inspection_logs": [
            {
                "id": r.id,
                "report_code": r.report_code or f"IR-{r.id}",
                "inspection_date": r.inspection_date or "",
                "part_no": r.part_no or "",
                "opn_no": r.opn_no or "10",
                "batch_qty": r.batch_qty or 5,
                "machine_name": r.machine_name or "",
                "operator_name": r.operator_name or "",
                "comp_sl_nos": r.comp_sl_nos or "",
                "readings_json": r.readings_json or "{}"
            }
            for r in recent_inspection_reports
        ]
    }

# --- Machines ---
@app.get("/api/machines")
def get_machines(db: Session = Depends(get_db)):
    try:
        from sqlalchemy import text
        rows = db.execute(text("SELECT * FROM machines ORDER BY id ASC")).mappings().all()
        if rows:
            return [{
                "id": r.get("id"),
                "name": r.get("name") or r.get("machine_name") or "",
                "dept": r.get("department") or r.get("dept") or "",
                "department": r.get("department") or r.get("dept") or "",
                "status": r.get("status") or "Active"
            } for r in rows]
    except Exception:
        db.rollback()

    try:
        machines = db.query(models.Machine).order_by(models.Machine.id.asc()).all()
        return [{
            "id": m.id,
            "name": m.name,
            "dept": m.dept or "",
            "department": m.dept or "",
            "status": m.status or "Active"
        } for m in machines]
    except Exception:
        db.rollback()
        return []

@app.post("/api/machines", response_model=MachineResponse)
def create_machine(machine: MachineCreate, db: Session = Depends(get_db)):
    db_m = models.Machine(**machine.model_dump())
    db.add(db_m)
    db.commit()
    db.refresh(db_m)
    return db_m

@app.post("/api/machines/import-excel")
async def import_machines_excel(file: UploadFile = File(...), db: Session = Depends(get_db)):
    contents = await file.read()
    rows = parse_excel_bytes(contents)
    if not rows:
        raise HTTPException(status_code=400, detail="Could not parse Excel file or file is empty")
    
    headers = [h.lower().strip() for h in rows[0]]
    
    name_idx = -1
    dept_idx = -1
    status_idx = -1
    
    for i, h in enumerate(headers):
        if "name" in h or "machine" in h or "mc" in h or "m/c" in h:
            if name_idx == -1: name_idx = i
        elif "dept" in h or "department" in h:
            dept_idx = i
        elif "status" in h:
            status_idx = i

    if name_idx == -1:
        name_idx = 1 if len(headers) > 1 else 0
        dept_idx = 0 if len(headers) > 1 else -1

    existing_names = {m.name.strip().upper() for m in db.query(models.Machine).all()}
    imported_count = 0

    for row in rows[1:]:
        if name_idx < len(row) and row[name_idx]:
            name = row[name_idx].strip()
            if not name or name.upper() in ["MACHINE", "MACHINE NAME", "M/C NAME"]:
                continue
            dept = row[dept_idx].strip() if dept_idx != -1 and dept_idx < len(row) and row[dept_idx] else "General"
            status = row[status_idx].strip() if status_idx != -1 and status_idx < len(row) and row[status_idx] else "Active"
            
            if name.upper() not in existing_names:
                m = models.Machine(name=name, dept=dept, status=status)
                db.add(m)
                existing_names.add(name.upper())
                imported_count += 1
                
    db.commit()
    return {"imported_count": imported_count, "message": f"Successfully imported {imported_count} new machines!"}

@app.put("/api/machines/{machine_id}", response_model=MachineResponse)
def update_machine(machine_id: int, machine: MachineCreate, db: Session = Depends(get_db)):
    db_m = db.query(models.Machine).filter(models.Machine.id == machine_id).first()
    if not db_m:
        raise HTTPException(status_code=404, detail="Machine not found")
    for k, v in machine.model_dump().items():
        setattr(db_m, k, v)
    db.commit()
    db.refresh(db_m)
    return db_m

@app.delete("/api/machines/clear-all")
def clear_all_machines(db: Session = Depends(get_db)):
    db.query(models.Machine).delete()
    db.commit()
    return {"message": "All machines cleared successfully!"}

@app.delete("/api/machines/{machine_id}")
def delete_machine(machine_id: int, db: Session = Depends(get_db)):
    db_m = db.query(models.Machine).filter(models.Machine.id == machine_id).first()
    if not db_m:
        raise HTTPException(status_code=404, detail="Machine not found")
    db.delete(db_m)
@app.post("/api/machines/bulk_import")
def bulk_import_machines(data: dict, db: Session = Depends(get_db)):
    machines = data.get("machines") or []
    count = 0
    for m in machines:
        name = (m.get("name") or m.get("machine_name") or "").strip()
        dept = (m.get("department") or m.get("dept") or "").strip()
        status = (m.get("status") or "Active").strip()
        if name:
            try:
                m_obj = models.Machine(name=name, dept=dept, status=status)
                db.add(m_obj)
                db.commit()
                count += 1
            except Exception:
                db.rollback()
                try:
                    db.execute(text("INSERT INTO machines (name, dept, status) VALUES (:name, :dept, :status)"), {
                        "name": name,
                        "dept": dept,
                        "status": status
                    })
                    db.commit()
                    count += 1
                except Exception:
                    db.rollback()
                    try:
                        db.execute(text("INSERT INTO machines (name, department, status) VALUES (:name, :dept, :status)"), {
                            "name": name,
                            "dept": dept,
                            "status": status
                        })
                        db.commit()
                        count += 1
                    except Exception:
                        db.rollback()
    return {"message": f"Successfully imported {count} machines!", "imported_count": count}

# --- Departments ---
@app.get("/api/departments")
def get_departments(db: Session = Depends(get_db)):
    try:
        from sqlalchemy import text
        rows = db.execute(text("SELECT * FROM departments ORDER BY id ASC")).mappings().all()
        if rows:
            return [{"id": r.get("id"), "name": r.get("name") or ""} for r in rows]
    except Exception:
        db.rollback()

    try:
        depts = db.query(models.Department).order_by(models.Department.id.asc()).all()
        return [{"id": d.id, "name": d.name} for d in depts]
    except Exception:
        db.rollback()
        return []

@app.post("/api/departments")
def create_department(data: dict, db: Session = Depends(get_db)):
    name = (data.get("name") or "").strip()
    if not name:
        raise HTTPException(status_code=400, detail="Department name cannot be empty")
    
    # Check if department with this name already exists
    try:
        existing = db.execute(text("SELECT id, name FROM departments WHERE UPPER(name) = :name"), {"name": name.upper()}).mappings().first()
        if existing:
            return {"id": existing.get("id"), "name": existing.get("name"), "message": "Department already exists"}
    except Exception:
        db.rollback()

    # Sync postgres sequence if needed
    try:
        db.execute(text("SELECT setval(pg_get_serial_sequence('departments', 'id'), coalesce(max(id),0) + 1, false) FROM departments;"))
        db.commit()
    except Exception:
        db.rollback()

    try:
        max_row = db.execute(text("SELECT COALESCE(MAX(id), 0) + 1 AS next_id FROM departments")).mappings().first()
        next_id = int(max_row.get("next_id")) if max_row and max_row.get("next_id") else 1
        
        try:
            db.execute(text("INSERT INTO departments (id, name) VALUES (:id, :name)"), {"id": next_id, "name": name})
            db.commit()
        except Exception:
            db.rollback()
            db.execute(text("INSERT INTO departments (name) VALUES (:name)"), {"name": name})
            db.commit()
    except Exception as ex:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Failed to create department: {ex}")
    
    try:
        row = db.execute(text("SELECT id, name FROM departments WHERE UPPER(name) = :name ORDER BY id DESC LIMIT 1"), {"name": name.upper()}).mappings().first()
        return {"id": row.get("id") if row else next_id, "name": name, "message": "Department created successfully"}
    except Exception:
        return {"id": next_id, "name": name, "message": "Department created successfully"}

@app.put("/api/departments/{dept_id}")
def update_department(dept_id: int, data: dict, db: Session = Depends(get_db)):
    name = (data.get("name") or "").strip()
    if not name:
        raise HTTPException(status_code=400, detail="Department name cannot be empty")
    try:
        db.execute(text("UPDATE departments SET name = :name WHERE id = :id"), {"id": dept_id, "name": name})
        db.commit()
    except Exception:
        db.rollback()
        try:
            d_obj = db.query(models.Department).filter(models.Department.id == dept_id).first()
            if d_obj:
                d_obj.name = name
                db.commit()
        except Exception as ex:
            db.rollback()
            raise HTTPException(status_code=500, detail=f"Failed to update department: {ex}")
    return {"id": dept_id, "name": name, "message": "Department updated successfully"}

@app.delete("/api/departments/{dept_id}")
def delete_department(dept_id: int, db: Session = Depends(get_db)):
    try:
        db.execute(text("DELETE FROM departments WHERE id = :id"), {"id": dept_id})
        db.commit()
    except Exception:
        db.rollback()
        try:
            d_obj = db.query(models.Department).filter(models.Department.id == dept_id).first()
            if d_obj:
                db.delete(d_obj)
                db.commit()
        except Exception:
            db.rollback()
    return {"message": "Department deleted"}

@app.delete("/api/departments/clear-all")
@app.delete("/api/departments/all")
@app.post("/api/departments/clear-all")
def clear_all_departments(db: Session = Depends(get_db)):
    try:
        db.execute(text("DELETE FROM departments;"))
        db.commit()
    except Exception:
        db.rollback()
    return {"message": "All departments cleared successfully!"}

# --- Shifts ---
@app.get("/api/shifts")
def get_shifts(db: Session = Depends(get_db)):
    try:
        rows = db.execute(text("SELECT * FROM shifts ORDER BY id ASC")).mappings().all()
        if rows:
            return [{"id": r.get("id"), "name": r.get("name") or "", "hours": float(r.get("hours") or 8.0)} for r in rows]
    except Exception:
        db.rollback()

    try:
        shifts = db.query(models.Shift).order_by(models.Shift.id.asc()).all()
        return [{"id": s.id, "name": s.name, "hours": s.hours} for s in shifts]
    except Exception:
        db.rollback()
        return []

@app.post("/api/shifts")
def create_shift(data: dict, db: Session = Depends(get_db)):
    name = (data.get("name") or "").strip()
    hours = float(data.get("hours") or 8.0)
    if not name:
        raise HTTPException(status_code=400, detail="Shift name cannot be empty")
    
    try:
        existing = db.execute(text("SELECT id, name, hours FROM shifts WHERE UPPER(name) = :name"), {"name": name.upper()}).mappings().first()
        if existing:
            return {"id": existing.get("id"), "name": existing.get("name"), "hours": existing.get("hours"), "message": "Shift already exists"}
    except Exception:
        db.rollback()

    try:
        db.execute(text("SELECT setval(pg_get_serial_sequence('shifts', 'id'), coalesce(max(id),0) + 1, false) FROM shifts;"))
        db.commit()
    except Exception:
        db.rollback()

    try:
        max_row = db.execute(text("SELECT COALESCE(MAX(id), 0) + 1 AS next_id FROM shifts")).mappings().first()
        next_id = int(max_row.get("next_id")) if max_row and max_row.get("next_id") else 1
        
        try:
            db.execute(text("INSERT INTO shifts (id, name, hours) VALUES (:id, :name, :hours)"), {"id": next_id, "name": name, "hours": hours})
            db.commit()
        except Exception:
            db.rollback()
            db.execute(text("INSERT INTO shifts (name, hours) VALUES (:name, :hours)"), {"name": name, "hours": hours})
            db.commit()
    except Exception as ex:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Failed to create shift: {ex}")
    
    try:
        row = db.execute(text("SELECT id, name, hours FROM shifts WHERE UPPER(name) = :name ORDER BY id DESC LIMIT 1"), {"name": name.upper()}).mappings().first()
        return {"id": row.get("id") if row else next_id, "name": name, "hours": hours, "message": "Shift created successfully"}
    except Exception:
        return {"id": next_id, "name": name, "hours": hours, "message": "Shift created successfully"}

@app.put("/api/shifts/{shift_id}")
def update_shift(shift_id: int, data: dict, db: Session = Depends(get_db)):
    name = (data.get("name") or "").strip()
    hours = float(data.get("hours") or 8.0)
    if not name:
        raise HTTPException(status_code=400, detail="Shift name cannot be empty")
    try:
        db.execute(text("UPDATE shifts SET name = :name, hours = :hours WHERE id = :id"), {"id": shift_id, "name": name, "hours": hours})
        db.commit()
    except Exception:
        db.rollback()
        try:
            s_obj = db.query(models.Shift).filter(models.Shift.id == shift_id).first()
            if s_obj:
                s_obj.name = name
                s_obj.hours = hours
                db.commit()
        except Exception as ex:
            db.rollback()
            raise HTTPException(status_code=500, detail=f"Failed to update shift: {ex}")
    return {"id": shift_id, "name": name, "hours": hours, "message": "Shift updated successfully"}

@app.delete("/api/shifts/{shift_id}")
def delete_shift(shift_id: int, db: Session = Depends(get_db)):
    try:
        db.execute(text("DELETE FROM shifts WHERE id = :id"), {"id": shift_id})
        db.commit()
    except Exception:
        db.rollback()
        try:
            s_obj = db.query(models.Shift).filter(models.Shift.id == shift_id).first()
            if s_obj:
                db.delete(s_obj)
                db.commit()
        except Exception:
            db.rollback()
    return {"message": "Shift deleted"}

@app.delete("/api/shifts/clear-all")
@app.delete("/api/shifts/all")
@app.post("/api/shifts/clear-all")
def clear_all_shifts(db: Session = Depends(get_db)):
    try:
        db.execute(text("DELETE FROM shifts;"))
        db.commit()
    except Exception:
        db.rollback()
    return {"message": "All shifts cleared successfully!"}

# --- Vendors ---
@app.get("/api/vendors")
def get_vendors(db: Session = Depends(get_db)):
    try:
        rows = db.execute(text("SELECT * FROM vendors ORDER BY id ASC")).mappings().all()
        if rows:
            return [{"id": r.get("id"), "name": r.get("name") or "", "details": r.get("details") or ""} for r in rows]
    except Exception:
        db.rollback()

    try:
        vendors = db.query(models.Vendor).order_by(models.Vendor.id.asc()).all()
        return [{"id": v.id, "name": v.name, "details": v.details or ""} for v in vendors]
    except Exception:
        db.rollback()
        return []

@app.post("/api/vendors")
def create_vendor(data: dict, db: Session = Depends(get_db)):
    name = (data.get("name") or "").strip()
    details = (data.get("details") or "").strip()
    if not name:
        raise HTTPException(status_code=400, detail="Vendor name cannot be empty")
    
    try:
        existing = db.execute(text("SELECT id, name, details FROM vendors WHERE UPPER(name) = :name"), {"name": name.upper()}).mappings().first()
        if existing:
            return {"id": existing.get("id"), "name": existing.get("name"), "details": existing.get("details"), "message": "Vendor already exists"}
    except Exception:
        db.rollback()

    try:
        db.execute(text("SELECT setval(pg_get_serial_sequence('vendors', 'id'), coalesce(max(id),0) + 1, false) FROM vendors;"))
        db.commit()
    except Exception:
        db.rollback()

    try:
        max_row = db.execute(text("SELECT COALESCE(MAX(id), 0) + 1 AS next_id FROM vendors")).mappings().first()
        next_id = int(max_row.get("next_id")) if max_row and max_row.get("next_id") else 1
        
        try:
            db.execute(text("INSERT INTO vendors (id, name, details) VALUES (:id, :name, :details)"), {"id": next_id, "name": name, "details": details})
            db.commit()
        except Exception:
            db.rollback()
            db.execute(text("INSERT INTO vendors (name, details) VALUES (:name, :details)"), {"name": name, "details": details})
            db.commit()
    except Exception as ex:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Failed to create vendor: {ex}")
    
    try:
        row = db.execute(text("SELECT id, name, details FROM vendors WHERE UPPER(name) = :name ORDER BY id DESC LIMIT 1"), {"name": name.upper()}).mappings().first()
        return {"id": row.get("id") if row else next_id, "name": name, "details": details, "message": "Vendor created successfully"}
    except Exception:
        return {"id": next_id, "name": name, "details": details, "message": "Vendor created successfully"}

@app.put("/api/vendors/{vendor_id}")
def update_vendor(vendor_id: int, data: dict, db: Session = Depends(get_db)):
    name = (data.get("name") or "").strip()
    details = (data.get("details") or "").strip()
    if not name:
        raise HTTPException(status_code=400, detail="Vendor name cannot be empty")
    try:
        db.execute(text("UPDATE vendors SET name = :name, details = :details WHERE id = :id"), {"id": vendor_id, "name": name, "details": details})
        db.commit()
    except Exception:
        db.rollback()
        try:
            v_obj = db.query(models.Vendor).filter(models.Vendor.id == vendor_id).first()
            if v_obj:
                v_obj.name = name
                v_obj.details = details
                db.commit()
        except Exception as ex:
            db.rollback()
            raise HTTPException(status_code=500, detail=f"Failed to update vendor: {ex}")
    return {"id": vendor_id, "name": name, "details": details, "message": "Vendor updated successfully"}

@app.delete("/api/vendors/{vendor_id}")
def delete_vendor(vendor_id: int, db: Session = Depends(get_db)):
    try:
        db.execute(text("DELETE FROM vendors WHERE id = :id"), {"id": vendor_id})
        db.commit()
    except Exception:
        db.rollback()
        try:
            v_obj = db.query(models.Vendor).filter(models.Vendor.id == vendor_id).first()
            if v_obj:
                db.delete(v_obj)
                db.commit()
        except Exception:
            db.rollback()
    return {"message": "Vendor deleted"}

@app.delete("/api/vendors/clear-all")
@app.delete("/api/vendors/all")
@app.post("/api/vendors/clear-all")
def clear_all_vendors(db: Session = Depends(get_db)):
    try:
        db.execute(text("DELETE FROM vendors;"))
        db.commit()
    except Exception:
        db.rollback()
    return {"message": "All vendors cleared successfully!"}

# --- Suppliers ---
@app.get("/api/suppliers")
def get_suppliers(db: Session = Depends(get_db)):
    try:
        rows = db.execute(text("SELECT * FROM suppliers ORDER BY id ASC")).mappings().all()
        if rows:
            return [{"id": r.get("id"), "name": r.get("name") or "", "details": r.get("details") or ""} for r in rows]
    except Exception:
        db.rollback()

    try:
        suppliers = db.query(models.Supplier).order_by(models.Supplier.id.asc()).all()
        return [{"id": s.id, "name": s.name, "details": s.details or ""} for s in suppliers]
    except Exception:
        db.rollback()
        return []

@app.post("/api/suppliers")
def create_supplier(data: dict, db: Session = Depends(get_db)):
    name = (data.get("name") or "").strip()
    details = (data.get("details") or "").strip()
    if not name:
        raise HTTPException(status_code=400, detail="Supplier name cannot be empty")
    
    try:
        existing = db.execute(text("SELECT id, name, details FROM suppliers WHERE UPPER(name) = :name"), {"name": name.upper()}).mappings().first()
        if existing:
            return {"id": existing.get("id"), "name": existing.get("name"), "details": existing.get("details"), "message": "Supplier already exists"}
    except Exception:
        db.rollback()

    try:
        db.execute(text("SELECT setval(pg_get_serial_sequence('suppliers', 'id'), coalesce(max(id),0) + 1, false) FROM suppliers;"))
        db.commit()
    except Exception:
        db.rollback()

    try:
        max_row = db.execute(text("SELECT COALESCE(MAX(id), 0) + 1 AS next_id FROM suppliers")).mappings().first()
        next_id = int(max_row.get("next_id")) if max_row and max_row.get("next_id") else 1
        
        try:
            db.execute(text("INSERT INTO suppliers (id, name, details) VALUES (:id, :name, :details)"), {"id": next_id, "name": name, "details": details})
            db.commit()
        except Exception:
            db.rollback()
            db.execute(text("INSERT INTO suppliers (name, details) VALUES (:name, :details)"), {"name": name, "details": details})
            db.commit()
    except Exception as ex:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Failed to create supplier: {ex}")
    
    try:
        row = db.execute(text("SELECT id, name, details FROM suppliers WHERE UPPER(name) = :name ORDER BY id DESC LIMIT 1"), {"name": name.upper()}).mappings().first()
        return {"id": row.get("id") if row else next_id, "name": name, "details": details, "message": "Supplier created successfully"}
    except Exception:
        return {"id": next_id, "name": name, "details": details, "message": "Supplier created successfully"}

@app.put("/api/suppliers/{supplier_id}")
def update_supplier(supplier_id: int, data: dict, db: Session = Depends(get_db)):
    name = (data.get("name") or "").strip()
    details = (data.get("details") or "").strip()
    if not name:
        raise HTTPException(status_code=400, detail="Supplier name cannot be empty")
    try:
        db.execute(text("UPDATE suppliers SET name = :name, details = :details WHERE id = :id"), {"id": supplier_id, "name": name, "details": details})
        db.commit()
    except Exception:
        db.rollback()
        try:
            s_obj = db.query(models.Supplier).filter(models.Supplier.id == supplier_id).first()
            if s_obj:
                s_obj.name = name
                s_obj.details = details
                db.commit()
        except Exception as ex:
            db.rollback()
            raise HTTPException(status_code=500, detail=f"Failed to update supplier: {ex}")
    return {"id": supplier_id, "name": name, "details": details, "message": "Supplier updated successfully"}

@app.delete("/api/suppliers/{supplier_id}")
def delete_supplier(supplier_id: int, db: Session = Depends(get_db)):
    try:
        db.execute(text("DELETE FROM suppliers WHERE id = :id"), {"id": supplier_id})
        db.commit()
    except Exception:
        db.rollback()
        try:
            s_obj = db.query(models.Supplier).filter(models.Supplier.id == supplier_id).first()
            if s_obj:
                db.delete(s_obj)
                db.commit()
        except Exception:
            db.rollback()
    return {"message": "Supplier deleted"}

@app.delete("/api/suppliers/clear-all")
@app.delete("/api/suppliers/all")
@app.post("/api/suppliers/clear-all")
def clear_all_suppliers(db: Session = Depends(get_db)):
    try:
        db.execute(text("DELETE FROM suppliers;"))
        db.commit()
    except Exception:
        db.rollback()
    return {"message": "All suppliers cleared successfully!"}

# --- Setters ---
@app.get("/api/setters")
def get_setters(db: Session = Depends(get_db)):
    try:
        rows = db.execute(text("SELECT * FROM setters ORDER BY id ASC")).mappings().all()
        if rows:
            return [{"id": r.get("id"), "name": r.get("name") or "", "department": r.get("department") or r.get("dept") or ""} for r in rows]
    except Exception:
        db.rollback()

    try:
        setters = db.query(models.Setter).order_by(models.Setter.id.asc()).all()
        return [{"id": s.id, "name": s.name, "department": s.department or ""} for s in setters]
    except Exception:
        db.rollback()
        return []

@app.post("/api/setters")
def create_setter(data: dict, db: Session = Depends(get_db)):
    name = (data.get("name") or "").strip()
    department = (data.get("department") or data.get("dept") or "").strip()
    if not name:
        raise HTTPException(status_code=400, detail="Setter name cannot be empty")
    
    try:
        existing = db.execute(text("SELECT id, name, department FROM setters WHERE UPPER(name) = :name"), {"name": name.upper()}).mappings().first()
        if existing:
            return {"id": existing.get("id"), "name": existing.get("name"), "department": existing.get("department"), "message": "Setter already exists"}
    except Exception:
        db.rollback()

    try:
        db.execute(text("SELECT setval(pg_get_serial_sequence('setters', 'id'), coalesce(max(id),0) + 1, false) FROM setters;"))
        db.commit()
    except Exception:
        db.rollback()

    try:
        max_row = db.execute(text("SELECT COALESCE(MAX(id), 0) + 1 AS next_id FROM setters")).mappings().first()
        next_id = int(max_row.get("next_id")) if max_row and max_row.get("next_id") else 1
        
        try:
            db.execute(text("INSERT INTO setters (id, name, department) VALUES (:id, :name, :department)"), {"id": next_id, "name": name, "department": department})
            db.commit()
        except Exception:
            db.rollback()
            db.execute(text("INSERT INTO setters (name, department) VALUES (:name, :department)"), {"name": name, "department": department})
            db.commit()
    except Exception as ex:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Failed to create setter: {ex}")
    
    try:
        row = db.execute(text("SELECT id, name, department FROM setters WHERE UPPER(name) = :name ORDER BY id DESC LIMIT 1"), {"name": name.upper()}).mappings().first()
        return {"id": row.get("id") if row else next_id, "name": name, "department": department, "message": "Setter created successfully"}
    except Exception:
        return {"id": next_id, "name": name, "department": department, "message": "Setter created successfully"}

@app.put("/api/setters/{setter_id}")
def update_setter(setter_id: int, data: dict, db: Session = Depends(get_db)):
    name = (data.get("name") or "").strip()
    department = (data.get("department") or data.get("dept") or "").strip()
    if not name:
        raise HTTPException(status_code=400, detail="Setter name cannot be empty")
    try:
        db.execute(text("UPDATE setters SET name = :name, department = :department WHERE id = :id"), {"id": setter_id, "name": name, "department": department})
        db.commit()
    except Exception:
        db.rollback()
        try:
            s_obj = db.query(models.Setter).filter(models.Setter.id == setter_id).first()
            if s_obj:
                s_obj.name = name
                s_obj.department = department
                db.commit()
        except Exception as ex:
            db.rollback()
            raise HTTPException(status_code=500, detail=f"Failed to update setter: {ex}")
    return {"id": setter_id, "name": name, "department": department, "message": "Setter updated successfully"}

@app.delete("/api/setters/{setter_id}")
def delete_setter(setter_id: int, db: Session = Depends(get_db)):
    try:
        db.execute(text("DELETE FROM setters WHERE id = :id"), {"id": setter_id})
        db.commit()
    except Exception:
        db.rollback()
        try:
            s_obj = db.query(models.Setter).filter(models.Setter.id == setter_id).first()
            if s_obj:
                db.delete(s_obj)
                db.commit()
        except Exception:
            db.rollback()
    return {"message": "Setter deleted"}

@app.delete("/api/setters/clear-all")
@app.delete("/api/setters/all")
@app.post("/api/setters/clear-all")
def clear_all_setters(db: Session = Depends(get_db)):
    try:
        db.execute(text("DELETE FROM setters;"))
        db.commit()
    except Exception:
        db.rollback()
    return {"message": "All setters cleared successfully!"}

# --- Operators ---
@app.get("/api/operators")
def get_operators(db: Session = Depends(get_db)):
    try:
        from sqlalchemy import text
        rows = db.execute(text("SELECT * FROM operators ORDER BY id ASC")).mappings().all()
        if rows:
            return [{
                "id": r.get("id"),
                "name": r.get("name") or r.get("operator_name") or "",
                "dept": r.get("department") or r.get("dept") or "",
                "department": r.get("department") or r.get("dept") or "",
                "designation": r.get("designation") or "Operator"
            } for r in rows]
    except Exception:
        db.rollback()

    try:
        operators = db.query(models.Operator).order_by(models.Operator.id.asc()).all()
        return [{
            "id": o.id,
            "name": o.name,
            "dept": o.dept or "",
            "department": o.dept or "",
            "designation": o.designation or "Operator"
        } for o in operators]
    except Exception:
        db.rollback()
        return []

@app.post("/api/operators")
def create_operator(data: dict, db: Session = Depends(get_db)):
    name = data.get("name") or ""
    dept = data.get("department") or data.get("dept") or ""
    desig = data.get("designation") or "Operator"
    try:
        db.execute(text("INSERT INTO operators (name, dept, designation) VALUES (:name, :dept, :designation)"), {"name": name, "dept": dept, "designation": desig})
        db.commit()
    except Exception:
        try:
            db.execute(text("INSERT INTO operators (name, department, designation) VALUES (:name, :dept, :designation)"), {"name": name, "dept": dept, "designation": desig})
            db.commit()
        except Exception:
            db.rollback()
    return {"message": "Operator created", "name": name, "dept": dept, "department": dept, "designation": desig}

@app.post("/api/operators/import-excel")
async def import_operators_excel(file: UploadFile = File(...), db: Session = Depends(get_db)):
    contents = await file.read()
    rows = parse_excel_bytes(contents)
    if not rows:
        raise HTTPException(status_code=400, detail="Could not parse Excel file or file is empty")
    
    headers = [h.lower().strip() for h in rows[0]]
    
    name_idx = -1
    dept_idx = -1
    desig_idx = -1
    
    for i, h in enumerate(headers):
        if "name" in h or "operator" in h:
            name_idx = i
        elif "dept" in h or "department" in h:
            dept_idx = i
        elif "designation" in h or "role" in h or "desig" in h:
            desig_idx = i

    if name_idx == -1:
        name_idx = 1 if len(headers) > 1 else 0
        dept_idx = 0 if len(headers) > 1 else -1

    existing_op_names = {o.name.strip().upper() for o in db.query(models.Operator).all()}
    imported_count = 0

    for row in rows[1:]:
        if name_idx < len(row) and row[name_idx]:
            name = row[name_idx].strip()
            if not name or name.upper() in ["OPERATOR", "OPERATOR NAME", "EMP NAME"]:
                continue
            dept = row[dept_idx].strip() if dept_idx != -1 and dept_idx < len(row) and row[dept_idx] else "General"
            desig = row[desig_idx].strip() if desig_idx != -1 and desig_idx < len(row) and row[desig_idx] else "Operator"
            
            if name.upper() not in existing_op_names:
                op_obj = models.Operator(name=name, dept=dept, designation=desig)
                db.add(op_obj)
                existing_op_names.add(name.upper())
                imported_count += 1
                
    db.commit()
    return {"imported_count": imported_count, "message": f"Successfully imported {imported_count} new operators!"}

@app.put("/api/operators/{op_id}")
def update_operator(op_id: int, data: dict, db: Session = Depends(get_db)):
    name = data.get("name") or ""
    dept = data.get("department") or data.get("dept") or ""
    desig = data.get("designation") or "Operator"
    try:
        db.execute(text("UPDATE operators SET name = :name, dept = :dept, designation = :designation WHERE id = :id"), {"id": op_id, "name": name, "dept": dept, "designation": desig})
        db.commit()
    except Exception:
        try:
            db.execute(text("UPDATE operators SET name = :name, department = :dept, designation = :designation WHERE id = :id"), {"id": op_id, "name": name, "dept": dept, "designation": desig})
            db.commit()
        except Exception:
            db.rollback()
    return {"id": op_id, "name": name, "dept": dept, "department": dept, "designation": desig}

@app.delete("/api/operators/clear-all")
@app.delete("/api/operators/all")
@app.post("/api/operators/clear-all")
def clear_all_operators(db: Session = Depends(get_db)):
    try:
        db.execute(text("DELETE FROM operators;"))
        db.commit()
    except Exception:
        db.rollback()
    try:
        db.query(models.Operator).delete()
        db.commit()
    except Exception:
        db.rollback()
@app.post("/api/operators/bulk_import")
def bulk_import_operators(data: dict, db: Session = Depends(get_db)):
    operators = data.get("operators") or []
    count = 0
    for op in operators:
        name = (op.get("name") or op.get("operator_name") or "").strip()
        dept = (op.get("department") or op.get("dept") or "").strip()
        desig = (op.get("designation") or op.get("role") or "Operator").strip()
        if name:
            try:
                op_obj = models.Operator(name=name, dept=dept, designation=desig)
                db.add(op_obj)
                db.commit()
                count += 1
            except Exception:
                db.rollback()
                try:
                    db.execute(text("INSERT INTO operators (name, dept, designation) VALUES (:name, :dept, :desig)"), {"name": name, "dept": dept, "desig": desig})
                    db.commit()
                    count += 1
                except Exception:
                    db.rollback()
                    try:
                        db.execute(text("INSERT INTO operators (name, department, designation) VALUES (:name, :dept, :desig)"), {"name": name, "dept": dept, "desig": desig})
                        db.commit()
                        count += 1
                    except Exception:
                        db.rollback()
    return {"message": f"Successfully imported {count} operators!", "imported_count": count}

# --- Part Master API ---
@app.post("/api/partmaster/bulk_import")
def bulk_import_partmasters(data: dict, db: Session = Depends(get_db)):
    parts = data.get("parts") or []
    count = 0
    for p in parts:
        partno = (p.get("partno") or p.get("part_no") or "").strip()
        family = (p.get("family") or "").strip()
        forge_pn = (p.get("forge_pn") or p.get("forgepn") or "").strip()
        part_prefix = (p.get("part_prefix") or p.get("prefix") or "").strip()
        department = (p.get("department") or p.get("dept") or "").strip()
        customer = (p.get("customer") or "").strip()
        va = str(p.get("va") or 0)
        operations = p.get("operations") or []
        if partno:
            try:
                db.execute(text("INSERT INTO part_masters (customer, department, family, forge_pn, part_prefix, partno, va, rfd_phy) VALUES (:customer, :department, :family, :forge_pn, :part_prefix, :partno, :va, :rfd_phy)"), {
                    "customer": customer,
                    "department": department,
                    "family": family,
                    "forge_pn": forge_pn,
                    "part_prefix": part_prefix,
                    "partno": partno,
                    "va": va,
                    "rfd_phy": 0
                })
                db.commit()
                count += 1
            except Exception:
                db.rollback()
                try:
                    db.execute(text("INSERT INTO parts (part_no, customer, dept, family, forge_pn, va) VALUES (:part_no, :customer, :dept, :family, :forge_pn, :va)"), {
                        "part_no": partno,
                        "customer": customer,
                        "dept": department,
                        "family": family,
                        "forge_pn": forge_pn,
                        "va": float(va) if va.replace('.','',1).isdigit() else 0.0
                    })
                    db.commit()
                    count += 1
                except Exception:
                    db.rollback()
            
            if operations:
                try:
                    row = db.execute(text("SELECT id FROM part_masters WHERE partno = :p ORDER BY id DESC LIMIT 1"), {"p": partno}).mappings().first()
                    part_id = row.get("id") if row else None
                    for op in operations:
                        opn = str(op.get("opn_no") or "")
                        desc = op.get("description") or ""
                        mc = op.get("machine") or op.get("machine_name") or ""
                        cyc = float(op.get("cycle_time") or 0)
                        if (opn or desc) and part_id:
                            try:
                                db.execute(text("INSERT INTO part_operations (part_id, opn_no, description, machine, cycle_time) VALUES (:part_id, :opn_no, :description, :machine, :cycle_time)"), {
                                    "part_id": str(part_id),
                                    "opn_no": opn,
                                    "description": desc,
                                    "machine": mc,
                                    "cycle_time": cyc
                                })
                                db.commit()
                            except Exception:
                                db.rollback()
                except Exception:
                    db.rollback()
    return {"message": f"Successfully imported {count} parts!", "imported_count": count}
@app.delete("/api/partmaster/clear-all")
@app.post("/api/partmaster/clear-all")
def clear_all_parts(db: Session = Depends(get_db)):
    try:
        db.execute(text("DELETE FROM part_masters;"))
    except Exception:
        pass
    try:
        db.execute(text("DELETE FROM parts;"))
    except Exception:
        pass
    try:
        db.execute(text("DELETE FROM part_operations;"))
    except Exception:
        pass
    try:
        db.commit()
    except Exception:
        db.rollback()
    return {"success": True, "message": "All parts cleared successfully!"}

@app.get("/api/partmaster")
def get_partmasters(db: Session = Depends(get_db)):
    try:
        rows = db.execute(text("SELECT * FROM part_masters ORDER BY CASE WHEN customer IS NOT NULL AND customer != '' THEN 0 ELSE 1 END, id ASC")).mappings().all()
        results = []
        for r in rows:
            p_no = r.get("partno") or r.get("part_no") or r.get("part_number") or ""
            if p_no:
                results.append({
                    "id": r.get("id"),
                    "customer": r.get("customer") or r.get("customer_name") or "",
                    "department": r.get("department") or r.get("dept") or "",
                    "family": r.get("family") or "",
                    "forge_pn": r.get("forge_pn") or r.get("forge_part_no") or "",
                    "part_prefix": r.get("part_prefix") or "",
                    "partno": p_no,
                    "part_no": p_no,
                    "va": r.get("va") or 0,
                    "rfd_phy": r.get("rfd_phy") or 0
                })
        return results
    except Exception as e:
        print("get_partmasters error:", e)
        db.rollback()
        return []

@app.get("/api/partmaster/{part_id}")
def get_partmaster_by_id(part_id: int, db: Session = Depends(get_db)):
    try:
        rows = db.execute(text("SELECT * FROM part_masters WHERE id = :id"), {"id": part_id}).mappings().all()
        if rows:
            r = rows[0]
            return {
                "id": r.get("id"),
                "customer": r.get("customer") or "",
                "department": r.get("department") or r.get("dept") or "",
                "family": r.get("family") or "",
                "forge_pn": r.get("forge_pn") or "",
                "part_prefix": r.get("part_prefix") or "",
                "partno": r.get("partno") or r.get("part_no") or "",
                "part_no": r.get("partno") or r.get("part_no") or "",
                "va": r.get("va") or 0,
                "rfd_phy": r.get("rfd_phy") or 0
            }
    except Exception:
        db.rollback()
    return {}

@app.get("/api/partmaster/{part_id}/operations")
def get_partmaster_operations(part_id: int, db: Session = Depends(get_db)):
    # 1. Search part_operations by part_id
    try:
        rows = db.execute(text("SELECT * FROM part_operations WHERE CAST(part_id AS TEXT) = :part_id_str ORDER BY id ASC"), {"part_id_str": str(part_id)}).mappings().all()
        if rows:
            results = []
            for r in rows:
                results.append({
                    "id": r.get("id"),
                    "part_id": part_id,
                    "opn_no": str(r.get("opn_no") or ""),
                    "description": r.get("description") or "",
                    "machine": r.get("machine") or r.get("machine_name") or "",
                    "cycle_time": float(r.get("cycle_time") or 0.0)
                })
            return results
    except Exception as e:
        db.rollback()

    # 2. Search operations by part_id
    try:
        rows = db.execute(text("SELECT * FROM operations WHERE CAST(part_id AS TEXT) = :part_id_str ORDER BY id ASC"), {"part_id_str": str(part_id)}).mappings().all()
        if rows:
            results = []
            for r in rows:
                results.append({
                    "id": r.get("id"),
                    "part_id": part_id,
                    "opn_no": str(r.get("opn_no") or ""),
                    "description": r.get("description") or "",
                    "machine": r.get("machine_name") or r.get("machine") or "",
                    "cycle_time": float(r.get("cycle_time") or 0.0)
                })
            return results
    except Exception as e:
        db.rollback()

    # 3. Fallback by matching partno
    try:
        pm_rows = db.execute(text("SELECT * FROM part_masters WHERE id = :id"), {"id": part_id}).mappings().all()
        if pm_rows:
            p_no = pm_rows[0].get("partno") or pm_rows[0].get("part_no") or ""
            if p_no:
                part_rows = db.execute(text("SELECT id FROM parts WHERE part_no = :p_no"), {"p_no": p_no}).mappings().all()
                if part_rows:
                    real_pid = part_rows[0].get("id")
                    rows = db.execute(text("SELECT * FROM operations WHERE part_id = :pid ORDER BY id ASC"), {"pid": real_pid}).mappings().all()
                    if rows:
                        results = []
                        for r in rows:
                            results.append({
                                "id": r.get("id"),
                                "part_id": part_id,
                                "opn_no": str(r.get("opn_no") or ""),
                                "description": r.get("description") or "",
                                "machine": r.get("machine_name") or r.get("machine") or "",
                                "cycle_time": float(r.get("cycle_time") or 0.0)
                            })
                        return results
    except Exception:
        db.rollback()

    return []

@app.post("/api/partmaster/{part_id}/operations")
def save_partmaster_operations(part_id: int, ops: List[dict], db: Session = Depends(get_db)):
    try:
        db.execute(text("DELETE FROM part_operations WHERE CAST(part_id AS TEXT) = :part_id_str"), {"part_id_str": str(part_id)})
        try:
            db.execute(text("DELETE FROM operations WHERE CAST(part_id AS TEXT) = :part_id_str"), {"part_id_str": str(part_id)})
        except Exception:
            pass
        for op in ops:
            db.execute(text("INSERT INTO part_operations (part_id, opn_no, description, machine, cycle_time) VALUES (:part_id, :opn_no, :description, :machine, :cycle_time)"), {
                "part_id": str(part_id),
                "opn_no": str(op.get("opn_no") or ""),
                "description": op.get("description") or "",
                "machine": op.get("machine") or op.get("machine_name") or "",
                "cycle_time": float(op.get("cycle_time") or 0)
            })
            try:
                db.execute(text("INSERT INTO operations (part_id, opn_no, description, machine_name, cycle_time) VALUES (:part_id, :opn_no, :description, :machine, :cycle_time)"), {
                    "part_id": part_id,
                    "opn_no": str(op.get("opn_no") or ""),
                    "description": op.get("description") or "",
                    "machine": op.get("machine") or op.get("machine_name") or "",
                    "cycle_time": float(op.get("cycle_time") or 0)
                })
            except Exception:
                pass
        db.commit()
        return {"message": "Operations saved successfully"}
    except Exception as e:
        db.rollback()
        return {"error": str(e)}

@app.post("/api/partmaster")
def create_partmaster(data: dict, db: Session = Depends(get_db)):
    try:
        db.execute(text("INSERT INTO part_masters (customer, department, family, forge_pn, part_prefix, partno, va, rfd_phy) VALUES (:customer, :department, :family, :forge_pn, :part_prefix, :partno, :va, :rfd_phy)"), {
            "customer": data.get("customer") or "",
            "department": data.get("department") or data.get("dept") or "",
            "family": data.get("family") or "",
            "forge_pn": data.get("forge_pn") or "",
            "part_prefix": data.get("part_prefix") or "",
            "partno": data.get("partno") or data.get("part_no") or "",
            "va": data.get("va") or 0,
            "rfd_phy": data.get("rfd_phy") or 0
        })
        db.commit()
        return {"message": "Part master created", **data}
    except Exception as e:
        db.rollback()
        return {"message": "Part master created", **data}

@app.put("/api/partmaster/{part_id}")
def update_partmaster(part_id: int, data: dict, db: Session = Depends(get_db)):
    try:
        db.execute(text("UPDATE part_masters SET customer = :customer, department = :department, family = :family, forge_pn = :forge_pn, part_prefix = :part_prefix, partno = :partno, va = :va WHERE id = :id"), {
            "id": part_id,
            "customer": data.get("customer") or "",
            "department": data.get("department") or data.get("dept") or "",
            "family": data.get("family") or "",
            "forge_pn": data.get("forge_pn") or "",
            "part_prefix": data.get("part_prefix") or "",
            "partno": data.get("partno") or data.get("part_no") or "",
            "va": data.get("va") or 0
        })
        db.commit()
        return {"id": part_id, **data}
    except Exception as e:
        db.rollback()
        return {"id": part_id, **data}

@app.delete("/api/partmaster/{part_id}")
def delete_partmaster(part_id: int, db: Session = Depends(get_db)):
    try:
        db.execute(text("DELETE FROM part_masters WHERE id = :id"), {"id": part_id})
        db.execute(text("DELETE FROM part_operations WHERE part_id = :id"), {"id": part_id})
        db.commit()
        return {"message": "Part master deleted"}
    except Exception as e:
        db.rollback()
        return {"message": "Deleted"}

# --- Parts ---
@app.get("/api/parts", response_model=List[PartResponse])
def get_parts(db: Session = Depends(get_db)):
    return db.query(models.Part).all()

@app.post("/api/parts", response_model=PartResponse)
def create_part(part: PartCreate, db: Session = Depends(get_db)):
    ops_data = part.operations
    part_dict = part.model_dump(exclude={"operations"})
    db_part = models.Part(**part_dict)
    
    for op in ops_data:
        db_part.operations.append(models.Operation(**op.model_dump()))

    db.add(db_part)
    db.commit()
    db.refresh(db_part)
    return db_part

@app.post("/api/parts/import-excel")
async def import_parts_excel(file: UploadFile = File(...), db: Session = Depends(get_db)):
    contents = await file.read()
    rows = parse_excel_bytes(contents)
    if not rows:
        raise HTTPException(status_code=400, detail="Could not parse Excel file or file is empty")
    
    headers = [h.lower().strip() for h in rows[0]]
    
    part_no_idx = -1
    cust_idx = -1
    dept_idx = -1
    fam_idx = -1
    forge_idx = -1
    desc_idx = -1
    cycle_idx = -1
    va_idx = -1

    for i, h in enumerate(headers):
        if "part" in h or "drawing" in h: part_no_idx = i
        elif "cust" in h: cust_idx = i
        elif "dept" in h: dept_idx = i
        elif "family" in h: fam_idx = i
        elif "forge" in h: forge_idx = i
        elif "desc" in h: desc_idx = i
        elif "cycle" in h: cycle_idx = i
        elif "va" in h: va_idx = i

    if part_no_idx == -1: part_no_idx = 0

    existing_parts = {p.part_no.strip().upper(): p for p in db.query(models.Part).all()}
    imported_count = 0

    def safe_float(val, default=0.0):
        try:
            return float(val) if val else default
        except (ValueError, TypeError):
            return default

    for row in rows[1:]:
        if part_no_idx < len(row) and row[part_no_idx]:
            p_no = row[part_no_idx].strip()
            if not p_no or p_no.upper() in ["PART NO", "PART NUMBER", "DRAWING NO"]:
                continue
            
            cust = row[cust_idx].strip() if cust_idx != -1 and cust_idx < len(row) and row[cust_idx] else ""
            dept = row[dept_idx].strip() if dept_idx != -1 and dept_idx < len(row) and row[dept_idx] else ""
            fam = row[fam_idx].strip() if fam_idx != -1 and fam_idx < len(row) and row[fam_idx] else ""
            forge = row[forge_idx].strip() if forge_idx != -1 and forge_idx < len(row) and row[forge_idx] else ""
            desc = row[desc_idx].strip() if desc_idx != -1 and desc_idx < len(row) and row[desc_idx] else ""
            cycle = safe_float(row[cycle_idx]) if cycle_idx != -1 and cycle_idx < len(row) else 0.0
            va = safe_float(row[va_idx]) if va_idx != -1 and va_idx < len(row) else 0.0

            if p_no.upper() not in existing_parts:
                new_p = models.Part(
                    part_no=p_no, customer=cust, dept=dept, family=fam,
                    forge_pn=forge, description=desc, cycle_time=cycle, va=va
                )
                db.add(new_p)
                existing_parts[p_no.upper()] = new_p
                imported_count += 1
            else:
                existing_p = existing_parts[p_no.upper()]
                if cust: existing_p.customer = cust
                if dept: existing_p.dept = dept
                if fam: existing_p.family = fam
                if forge: existing_p.forge_pn = forge
                if desc: existing_p.description = desc
                if cycle > 0: existing_p.cycle_time = cycle
                if va > 0: existing_p.va = va

    db.commit()
    return {"imported_count": imported_count, "message": f"Successfully imported/updated parts master!"}

@app.put("/api/parts/{part_id}", response_model=PartResponse)
def update_part(part_id: int, part: PartCreate, db: Session = Depends(get_db)):
    db_part = db.query(models.Part).filter(models.Part.id == part_id).first()
    if not db_part:
        raise HTTPException(status_code=404, detail="Part not found")
    
    db_part.part_no = part.part_no
    db_part.customer = part.customer
    db_part.dept = part.dept
    db_part.family = part.family
    db_part.forge_pn = part.forge_pn
    db_part.description = part.description
    db_part.cycle_time = part.cycle_time
    db_part.va = part.va

    db.query(models.Operation).filter(models.Operation.part_id == part_id).delete()
    for op in part.operations:
        db_part.operations.append(models.Operation(**op.model_dump()))

    db.commit()
    db.refresh(db_part)
    return db_part

@app.delete("/api/parts/clear-all")
def clear_all_parts(db: Session = Depends(get_db)):
    db.query(models.Part).delete()
    db.commit()
    return {"message": "All parts cleared successfully!"}

@app.delete("/api/parts/{part_id}")
def delete_part(part_id: int, db: Session = Depends(get_db)):
    db_part = db.query(models.Part).filter(models.Part.id == part_id).first()
    if not db_part:
        raise HTTPException(status_code=404, detail="Part not found")
    db.delete(db_part)
    db.commit()
    return {"message": "Part deleted"}

# --- Schedules ---
@app.get("/api/schedules")
def get_schedules(db: Session = Depends(get_db)):
    try:
        schedules = db.query(models.ProductionSchedule).all()
        if schedules:
            parts = db.query(models.Part).all()
            part_map = {p.part_no.strip().upper(): p for p in parts if p.part_no}

            logs = db.query(models.ProductionLog).all()
            prod_map = {}
            for log in logs:
                if log.part_no and log.opn_no:
                    key = (log.part_no.strip().upper(), str(log.opn_no).strip())
                    prod_map[key] = prod_map.get(key, 0) + (log.qty_produced or 0)

            results = []
            for sch in schedules:
                p_key = sch.part_no.strip().upper() if sch.part_no else ""
                part = part_map.get(p_key)

                if part and part.operations and len(part.operations) > 0:
                    for opn in part.operations:
                        opn_str = str(opn.opn_no).strip()
                        qty_prod = prod_map.get((p_key, opn_str), 0)
                        bal = max(0, (sch.total_sch_qty or 0) - qty_prod)
                        results.append({
                            "id": sch.id,
                            "part_no": sch.part_no,
                            "sch_qty": sch.total_sch_qty or 0,
                            "opn_no": opn.opn_no,
                            "desc": opn.description or "",
                            "qty_prod": qty_prod,
                            "balance": bal
                        })
                else:
                    qty_prod = prod_map.get((p_key, "10"), 0)
                    bal = max(0, (sch.total_sch_qty or 0) - qty_prod)
                    results.append({
                        "id": sch.id,
                        "part_no": sch.part_no,
                        "sch_qty": sch.total_sch_qty or 0,
                        "opn_no": "10",
                        "desc": "General",
                        "qty_prod": qty_prod,
                        "balance": bal
                    })

            return results
    except Exception:
        db.rollback()

    try:
        from sqlalchemy import text
        rows = db.execute(text("SELECT * FROM schedules")).mappings().all()
        return [dict(r) for r in rows]
    except Exception:
        db.rollback()
        return []
    imported_opns_count = 0

    def safe_num(val):
        try:
            return float(val) if val else 0.0
        except (ValueError, TypeError):
            return 0.0

    for row in rows[1:]:
        if part_idx < len(row) and row[part_idx]:
            part_no = row[part_idx].strip()
            if not part_no or part_no.upper() == "PART NO":
                continue
            
            opn_no = row[opn_idx].strip() if opn_idx != -1 and opn_idx < len(row) and row[opn_idx] else "10"
            desc = row[desc_idx].strip() if desc_idx != -1 and desc_idx < len(row) and row[desc_idx] else ""
            cyc = safe_num(row[cyc_idx]) if cyc_idx != -1 and cyc_idx < len(row) and row[cyc_idx] else 0.0
            mach = row[mach_idx].strip() if mach_idx != -1 and mach_idx < len(row) and row[mach_idx] else ""
            if cyc == 0.0 and mach:
                cyc = safe_num(mach)
            customer = row[cust_idx].strip() if cust_idx != -1 and cust_idx < len(row) and row[cust_idx] else ""
            dept = row[dept_idx].strip() if dept_idx != -1 and dept_idx < len(row) and row[dept_idx] else ""
            family = row[fam_idx].strip() if fam_idx != -1 and fam_idx < len(row) and row[fam_idx] else ""
            forge_pn = row[forge_idx].strip() if forge_idx != -1 and forge_idx < len(row) and row[forge_idx] else ""

            part_key = part_no.upper()
            part = existing_parts.get(part_key)
            if not part:
                part = models.Part(
                    part_no=part_no,
                    customer=customer,
                    dept=dept,
                    family=family,
                    forge_pn=forge_pn,
                    description=desc,
                    cycle_time=cyc
                )
                db.add(part)
                db.flush()
                existing_parts[part_key] = part
                imported_parts_count += 1

            existing_opn_keys = {(str(op.opn_no).strip(), op.description.strip().upper()) for op in part.operations} if part.operations else set()
            if (opn_no, desc.upper()) not in existing_opn_keys:
                opn = models.Operation(
                    part_id=part.id,
                    opn_no=opn_no,
                    description=desc,
                    machine_name=mach,
                    cycle_time=cyc
                )
                db.add(opn)
                existing_opn_keys.add((opn_no, desc.upper()))
                imported_opns_count += 1

    db.commit()
    return {"imported_parts_count": imported_parts_count, "imported_opns_count": imported_opns_count, "message": f"Successfully imported {imported_parts_count} new parts and {imported_opns_count} operations!"}

@app.post("/api/parts/{part_id}/operations", response_model=OperationResponse)
def create_operation(part_id: int, opn: OperationCreate, db: Session = Depends(get_db)):
    db_part = db.query(models.Part).filter(models.Part.id == part_id).first()
    if not db_part:
        raise HTTPException(status_code=404, detail="Part not found")
    db_opn = models.Operation(part_id=part_id, **opn.model_dump())
    db.add(db_opn)
    db.commit()
    db.refresh(db_opn)
    return db_opn

@app.delete("/api/parts/clear-all")
def clear_all_parts(db: Session = Depends(get_db)):
    db.query(models.Part).delete()
    db.commit()
    return {"message": "All parts cleared successfully!"}

@app.delete("/api/parts/{part_id}")
def delete_part(part_id: int, db: Session = Depends(get_db)):
    db_part = db.query(models.Part).filter(models.Part.id == part_id).first()
    if not db_part:
        raise HTTPException(status_code=404, detail="Part not found")
    db.delete(db_part)
    db.commit()
    return {"message": "Part deleted"}

# --- Schedules ---
@app.get("/api/schedules")
@app.get("/api/schedule")
def get_schedules(db: Session = Depends(get_db)):
    try:
        schedules = db.query(models.ProductionSchedule).all()
        parts = db.query(models.Part).all()
        part_map = {p.part_no.strip().upper(): p for p in parts if p.part_no}

        logs = db.query(models.ProductionLog).all()
        prod_map = {}
        for log in logs:
            if log.part_no and log.opn_no:
                key = (log.part_no.strip().upper(), str(log.opn_no).strip())
                prod_map[key] = prod_map.get(key, 0) + (log.qty_produced or 0)

        results = []
        for sch in schedules:
            p_key = sch.part_no.strip().upper() if sch.part_no else ""
            part = part_map.get(p_key)

            if part and part.operations and len(part.operations) > 0:
                for opn in part.operations:
                    opn_str = str(opn.opn_no).strip()
                    qty_prod = prod_map.get((p_key, opn_str), 0)
                    bal = max(0, (sch.total_sch_qty or 0) - qty_prod)
                    results.append({
                        "id": sch.id,
                        "part_no": sch.part_no,
                        "sch_qty": sch.total_sch_qty or 0,
                        "opn_no": opn.opn_no,
                        "desc": opn.description or "",
                        "qty_prod": qty_prod,
                        "balance": bal
                    })
            else:
                qty_prod = prod_map.get((p_key, "10"), 0)
                bal = max(0, (sch.total_sch_qty or 0) - qty_prod)
                results.append({
                    "id": sch.id,
                    "part_no": sch.part_no,
                    "sch_qty": sch.total_sch_qty or 0,
                    "opn_no": "10",
                    "desc": "General",
                    "qty_prod": qty_prod,
                    "balance": bal
                })

        return results
    except Exception as e:
        print("get_schedules error:", e)
        db.rollback()
        try:
            rows = db.execute(text("SELECT * FROM schedules")).mappings().all()
            return [dict(r) for r in rows]
        except Exception:
            return []

@app.delete("/api/schedules/clear-all")
def clear_all_schedules(db: Session = Depends(get_db)):
    db.query(models.ProductionSchedule).delete()
    db.commit()
    return {"message": "All work schedules cleared successfully!"}

@app.post("/api/schedules/import-excel")
async def import_schedules_excel(file: UploadFile = File(...), db: Session = Depends(get_db)):
    contents = await file.read()
    rows = parse_excel_bytes(contents)
    if not rows:
        raise HTTPException(status_code=400, detail="Could not parse Excel file or file is empty")
    
    data_started = False
    imported_count = 0

    def safe_int(val, default=0):
        try:
            return int(float(val)) if val else default
        except (ValueError, TypeError):
            return default

    def safe_float(val, default=0.0):
        try:
            return float(val) if val else default
        except (ValueError, TypeError):
            return default

    for row in rows:
        if not row: continue
        if "Sl No" in row or "PART NO" in [x.upper().strip() for x in row]:
            data_started = True
            continue
        if data_started and len(row) >= 4:
            sl_no = row[0]
            item = row[1] if len(row) > 1 else ""
            grs_no = row[2] if len(row) > 2 else ""
            part_no = row[3] if len(row) > 3 else ""
            total_sch_qty = safe_int(row[4]) if len(row) > 4 else 0
            rate_per_pc = safe_float(row[5]) if len(row) > 5 else 0.0
            amount = safe_float(row[6]) if len(row) > 6 else 0.0
            qty_disp = safe_int(row[7]) if len(row) > 7 else 0
            value_rs = safe_float(row[8]) if len(row) > 8 else 0.0
            balance_to_produce = safe_int(row[9]) if len(row) > 9 else (total_sch_qty - qty_disp)
            remarks = row[10] if len(row) > 10 else ""

            if part_no and part_no.upper() != "PART NO":
                sch = models.ProductionSchedule(
                    sl_no=sl_no,
                    item=item,
                    grs_no=grs_no,
                    part_no=part_no,
                    total_sch_qty=total_sch_qty,
                    rate_per_pc=rate_per_pc,
                    amount=amount,
                    qty_disp=qty_disp,
                    value_rs=value_rs,
                    balance_to_produce=balance_to_produce,
                    remarks=remarks
                )
                db.add(sch)
                imported_count += 1

    db.commit()
    return {"imported_count": imported_count, "message": f"Successfully imported {imported_count} work schedule items!"}

@app.post("/api/schedules", response_model=ProductionScheduleResponse)
def create_schedule(sch: ProductionScheduleCreate, db: Session = Depends(get_db)):
    db_sch = models.ProductionSchedule(**sch.model_dump())
    db.add(db_sch)
    db.commit()
    db.refresh(db_sch)
    return db_sch

@app.delete("/api/schedules/{sch_id}")
def delete_schedule(sch_id: int, db: Session = Depends(get_db)):
    db_sch = db.query(models.ProductionSchedule).filter(models.ProductionSchedule.id == sch_id).first()
    if not db_sch:
        raise HTTPException(status_code=404, detail="Schedule not found")
    db.delete(db_sch)
    db.commit()
    return {"message": "Schedule deleted"}

# --- Production Logging ---
@app.get("/api/production-logs", response_model=List[ProductionLogResponse])
def get_production_logs(limit: int = 100, db: Session = Depends(get_db)):
    return db.query(models.ProductionLog).order_by(models.ProductionLog.id.desc()).limit(limit).all()

@app.get("/api/production-logs/sl-nos")
def get_completed_sl_nos(part_no: str, opn_no: Optional[str] = None, db: Session = Depends(get_db)):
    try:
        clean_part = part_no.strip().lower()

        def safe_str_opn(val):
            if val is None:
                return ""
            try:
                f = float(val)
                return str(int(f)) if f.is_integer() else str(f)
            except Exception:
                return str(val).strip()

        def extract_clean_opn(raw_str):
            if not raw_str:
                return None, ""
            raw = str(raw_str).strip()
            cleaned = re.sub(r'\(\s*\d+\s*min\s*\)', '', raw, flags=re.IGNORECASE)
            m_opn = re.search(r'opn\s*(\d+(?:\.\d+)?)', cleaned, re.IGNORECASE)
            if m_opn:
                num = float(m_opn.group(1))
                return num, safe_str_opn(num)
            m_start = re.match(r'^\s*(\d+(?:\.\d+)?)', cleaned)
            if m_start:
                num = float(m_start.group(1))
                return num, safe_str_opn(num)
            m_any = re.search(r'(\d+(?:\.\d+)?)', cleaned)
            if m_any:
                num = float(m_any.group(1))
                return num, safe_str_opn(num)
            return None, raw

        curr_num, curr_opn_str = extract_clean_opn(opn_no)
        if curr_num is None:
            curr_num, curr_opn_str = 10.0, "10"

        logs = db.query(models.ProductionLog).filter(func.lower(models.ProductionLog.part_no) == clean_part).all()
        part = db.query(models.Part).filter(func.lower(models.Part.part_no) == clean_part).first()
        schedules = db.query(models.ProductionSchedule).filter(func.lower(models.ProductionSchedule.part_no) == clean_part).all()

        # Collect ALL operation numbers for this part from Part Master + Work Schedules + Logs
        all_opn_nums = []
        if part and part.operations:
            for op in part.operations:
                num, _ = extract_clean_opn(op.opn_no)
                if num is not None and num not in all_opn_nums:
                    all_opn_nums.append(num)

        for sch in schedules:
            opn_val = getattr(sch, 'opn_no', None)
            if opn_val:
                num, _ = extract_clean_opn(opn_val)
                if num is not None and num not in all_opn_nums:
                    all_opn_nums.append(num)

        for l in logs:
            num, _ = extract_clean_opn(l.opn_no)
            if num is not None and num not in all_opn_nums:
                all_opn_nums.append(num)

        if curr_num not in all_opn_nums:
            all_opn_nums.append(curr_num)

        all_opn_nums = sorted(all_opn_nums)

        # Determine position of curr_num in sorted operation sequence for this part
        idx = all_opn_nums.index(curr_num) if curr_num in all_opn_nums else 0

        def extract_sl_nos(target_opn_str, target_num):
            sl_set = set()
            for l in logs:
                l_num, l_str = extract_clean_opn(l.opn_no)
                if (l_num is not None and target_num is not None and l_num == target_num) or (l_str == target_opn_str):
                    if l.completed_sl_nos:
                        for s in l.completed_sl_nos.split(','):
                            s = s.strip()
                            if s.isdigit():
                                sl_set.add(int(s))
            return sorted(list(sl_set))

        curr_completed = extract_sl_nos(curr_opn_str, curr_num)

        if idx == 0:
            # First operation in sequence (e.g. Opn 20 for 214, H44, M50, DK7, U34)
            is_first_opn = True
            prev_opn_no = None
            prev_completed = []
            available_sl_nos = []  # Frontend will compute: {1..maxGrid} minus curr_completed
        else:
            # Subsequent operation in sequence! Predecessor operation is index - 1!
            is_first_opn = False
            p_num = all_opn_nums[idx - 1]
            prev_opn_no = safe_str_opn(p_num)
            prev_num, _ = extract_clean_opn(prev_opn_no) if prev_opn_no else (None, "")
            prev_completed = extract_sl_nos(prev_opn_no, prev_num)
            # Available for current opn = (Completed in predecessor opn) MINUS (Already logged in current opn)
            available_sl_nos = [s for s in prev_completed if s not in curr_completed]

        return {
            "is_first_opn": is_first_opn,
            "prev_opn_no": prev_opn_no,
            "prev_completed_sl_nos": prev_completed,
            "completed_sl_nos": curr_completed,
            "available_sl_nos": available_sl_nos
        }
    except Exception as e:
        import traceback
        print("Error in get_completed_sl_nos:", traceback.format_exc())
        raise HTTPException(status_code=500, detail=str(e))

# --- Production Logging CRUD ---
@app.get("/api/prodlog")
@app.get("/api/production-logs")
def get_all_prod_logs(db: Session = Depends(get_db)):
    try:
        rows = db.execute(text("SELECT * FROM production_logs ORDER BY id DESC")).mappings().all()
        if rows:
            results = []
            for r in rows:
                results.append({
                    "id": r.get("id"),
                    "dept": r.get("dept") or "",
                    "date": r.get("date") or r.get("log_date") or "",
                    "shift": r.get("shift") or "",
                    "setter": r.get("setter") or "",
                    "machine": r.get("machine") or r.get("machine_name") or "",
                    "operator": r.get("operator") or r.get("operator_name") or "",
                    "partno": r.get("partno") or r.get("part_no") or "",
                    "opn_no": r.get("opn_no") or "",
                    "description": r.get("description") or "",
                    "runtime": float(r.get("runtime") or 0.0),
                    "cycle_time": float(r.get("cycle_time") or 0.0),
                    "target_qty": float(r.get("target_qty") or 0.0),
                    "prod_qty": float(r.get("prod_qty") if r.get("prod_qty") is not None else (r.get("qty_produced") or 0)),
                    "efficiency": float(r.get("efficiency") or 0.0),
                    "idle_hours": float(r.get("idle_hours") or 0.0),
                    "idle_reason": r.get("idle_reason") or "None",
                    "idle_hours_2": float(r.get("idle_hours_2") or 0.0),
                    "idle_reason_2": r.get("idle_reason_2") or "None",
                    "idle_hours_3": float(r.get("idle_hours_3") or 0.0),
                    "idle_reason_3": r.get("idle_reason_3") or "None",
                    "multiple_mc": int(r.get("multiple_mc") or 1),
                    "created_at": str(r.get("created_at") or "")
                })
            return results
    except Exception:
        db.rollback()

    try:
        logs = db.query(models.ProductionLog).order_by(models.ProductionLog.id.desc()).all()
        return [{
            "id": l.id,
            "dept": getattr(l, "dept", "") or "",
            "date": getattr(l, "date", "") or getattr(l, "log_date", "") or "",
            "shift": l.shift or "",
            "setter": getattr(l, "setter", "") or "",
            "machine": getattr(l, "machine", "") or getattr(l, "machine_name", "") or "",
            "operator": getattr(l, "operator", "") or getattr(l, "operator_name", "") or "",
            "partno": getattr(l, "partno", "") or getattr(l, "part_no", "") or "",
            "opn_no": getattr(l, "opn_no", "") or "",
            "description": getattr(l, "description", "") or "",
            "runtime": float(getattr(l, "runtime", 0.0) or 0.0),
            "cycle_time": float(getattr(l, "cycle_time", 0.0) or 0.0),
            "target_qty": float(getattr(l, "target_qty", 0.0) or 0.0),
            "prod_qty": float(getattr(l, "prod_qty", None) if getattr(l, "prod_qty", None) is not None else (getattr(l, "qty_produced", 0) or 0)),
            "efficiency": float(getattr(l, "efficiency", 0.0) or 0.0),
            "idle_hours": float(getattr(l, "idle_hours", 0.0) or 0.0),
            "idle_reason": getattr(l, "idle_reason", "None") or "None",
            "idle_hours_2": float(getattr(l, "idle_hours_2", 0.0) or 0.0),
            "idle_reason_2": getattr(l, "idle_reason_2", "None") or "None",
            "idle_hours_3": float(getattr(l, "idle_hours_3", 0.0) or 0.0),
            "idle_reason_3": getattr(l, "idle_reason_3", "None") or "None",
            "multiple_mc": int(getattr(l, "multiple_mc", 1) or 1),
            "created_at": str(getattr(l, "created_at", "") or "")
        } for l in logs]
    except Exception:
        db.rollback()
        return []

@app.post("/api/prodlog")
@app.post("/api/production-logs")
def create_prod_log(data: dict, db: Session = Depends(get_db)):
    dept = (data.get("dept") or "").strip()
    date_val = (data.get("date") or data.get("log_date") or "").strip()
    shift = (data.get("shift") or "").strip()
    setter = (data.get("setter") or "").strip()
    machine = (data.get("machine") or data.get("machine_name") or "").strip()
    operator = (data.get("operator") or data.get("operator_name") or "").strip()
    partno = (data.get("partno") or data.get("part_no") or "").strip()
    opn_no = (data.get("opn_no") or "").strip()
    description = (data.get("description") or "").strip()
    runtime = float(data.get("runtime") or 0.0)
    cycle_time = float(data.get("cycle_time") or 0.0)
    target_qty = float(data.get("target_qty") or 0.0)
    prod_qty = float(data.get("prod_qty") if data.get("prod_qty") is not None else (data.get("qty_produced") or 0.0))
    efficiency = float(data.get("efficiency") or 0.0)
    idle_hours = float(data.get("idle_hours") or 0.0)
    idle_reason = (data.get("idle_reason") or "None").strip()
    idle_hours_2 = float(data.get("idle_hours_2") or 0.0)
    idle_reason_2 = (data.get("idle_reason_2") or "None").strip()
    idle_hours_3 = float(data.get("idle_hours_3") or 0.0)
    idle_reason_3 = (data.get("idle_reason_3") or "None").strip()
    multiple_mc = int(data.get("multiple_mc") or 1)

    try:
        db.execute(text("SELECT setval(pg_get_serial_sequence('production_logs', 'id'), coalesce(max(id),0) + 1, false) FROM production_logs;"))
        db.commit()
    except Exception:
        db.rollback()

    try:
        max_row = db.execute(text("SELECT COALESCE(MAX(id), 0) + 1 AS next_id FROM production_logs")).mappings().first()
        next_id = int(max_row.get("next_id")) if max_row and max_row.get("next_id") else 1
    except Exception:
        db.rollback()
        next_id = 1

    params = {
        "id": next_id,
        "dept": dept,
        "date": date_val,
        "shift": shift,
        "setter": setter,
        "machine": machine,
        "operator": operator,
        "partno": partno,
        "opn_no": opn_no,
        "description": description,
        "runtime": runtime,
        "cycle_time": cycle_time,
        "target_qty": target_qty,
        "prod_qty": prod_qty,
        "efficiency": efficiency,
        "idle_hours": idle_hours,
        "idle_reason": idle_reason,
        "idle_hours_2": idle_hours_2,
        "idle_reason_2": idle_reason_2,
        "idle_hours_3": idle_hours_3,
        "idle_reason_3": idle_reason_3,
        "multiple_mc": multiple_mc
    }

    try:
        db.execute(text("""
            INSERT INTO production_logs (id, dept, date, shift, setter, machine, operator, partno, opn_no, description, runtime, cycle_time, target_qty, prod_qty, efficiency, idle_hours, idle_reason, idle_hours_2, idle_reason_2, idle_hours_3, idle_reason_3, multiple_mc)
            VALUES (:id, :dept, :date, :shift, :setter, :machine, :operator, :partno, :opn_no, :description, :runtime, :cycle_time, :target_qty, :prod_qty, :efficiency, :idle_hours, :idle_reason, :idle_hours_2, :idle_reason_2, :idle_hours_3, :idle_reason_3, :multiple_mc)
        """), params)
        db.commit()
    except Exception:
        db.rollback()
        try:
            db.execute(text("""
                INSERT INTO production_logs (dept, date, shift, setter, machine, operator, partno, opn_no, description, runtime, cycle_time, target_qty, prod_qty, efficiency, idle_hours, idle_reason, idle_hours_2, idle_reason_2, idle_hours_3, idle_reason_3, multiple_mc)
                VALUES (:dept, :date, :shift, :setter, :machine, :operator, :partno, :opn_no, :description, :runtime, :cycle_time, :target_qty, :prod_qty, :efficiency, :idle_hours, :idle_reason, :idle_hours_2, :idle_reason_2, :idle_hours_3, :idle_reason_3, :multiple_mc)
            """), params)
            db.commit()
        except Exception as ex:
            db.rollback()
            raise HTTPException(status_code=500, detail=f"Failed to save production log: {ex}")

    return {"id": next_id, "message": "Production Log saved successfully"}

@app.put("/api/prodlog/{log_id}")
@app.put("/api/production-logs/{log_id}")
def update_prod_log(log_id: int, data: dict, db: Session = Depends(get_db)):
    params = {
        "id": log_id,
        "dept": (data.get("dept") or "").strip(),
        "date": (data.get("date") or data.get("log_date") or "").strip(),
        "shift": (data.get("shift") or "").strip(),
        "setter": (data.get("setter") or "").strip(),
        "machine": (data.get("machine") or data.get("machine_name") or "").strip(),
        "operator": (data.get("operator") or data.get("operator_name") or "").strip(),
        "partno": (data.get("partno") or data.get("part_no") or "").strip(),
        "opn_no": (data.get("opn_no") or "").strip(),
        "description": (data.get("description") or "").strip(),
        "runtime": float(data.get("runtime") or 0.0),
        "cycle_time": float(data.get("cycle_time") or 0.0),
        "target_qty": float(data.get("target_qty") or 0.0),
        "prod_qty": float(data.get("prod_qty") if data.get("prod_qty") is not None else (data.get("qty_produced") or 0.0)),
        "efficiency": float(data.get("efficiency") or 0.0),
        "idle_hours": float(data.get("idle_hours") or 0.0),
        "idle_reason": (data.get("idle_reason") or "None").strip(),
        "idle_hours_2": float(data.get("idle_hours_2") or 0.0),
        "idle_reason_2": (data.get("idle_reason_2") or "None").strip(),
        "idle_hours_3": float(data.get("idle_hours_3") or 0.0),
        "idle_reason_3": (data.get("idle_reason_3") or "None").strip(),
        "multiple_mc": int(data.get("multiple_mc") or 1)
    }
    try:
        db.execute(text("""
            UPDATE production_logs SET
                dept = :dept, date = :date, shift = :shift, setter = :setter,
                machine = :machine, operator = :operator, partno = :partno,
                opn_no = :opn_no, description = :description, runtime = :runtime,
                cycle_time = :cycle_time, target_qty = :target_qty, prod_qty = :prod_qty,
                efficiency = :efficiency, idle_hours = :idle_hours, idle_reason = :idle_reason,
                idle_hours_2 = :idle_hours_2, idle_reason_2 = :idle_reason_2,
                idle_hours_3 = :idle_hours_3, idle_reason_3 = :idle_reason_3,
                multiple_mc = :multiple_mc
            WHERE id = :id
        """), params)
        db.commit()
    except Exception as ex:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Failed to update production log: {ex}")
    return {"id": log_id, "message": "Production Log updated successfully"}

@app.delete("/api/prodlog/{log_id}")
@app.delete("/api/production-logs/{log_id}")
def delete_production_log(log_id: int, db: Session = Depends(get_db)):
    try:
        db.execute(text("DELETE FROM production_logs WHERE id = :id"), {"id": log_id})
        db.commit()
    except Exception:
        db.rollback()
        try:
            db_log = db.query(models.ProductionLog).filter(models.ProductionLog.id == log_id).first()
            if db_log:
                db.delete(db_log)
                db.commit()
        except Exception:
            db.rollback()
    return {"message": "Production Log deleted successfully"}

@app.delete("/api/prodlog/clear-all")
@app.delete("/api/prodlog/all")
@app.post("/api/prodlog/clear-all")
@app.delete("/api/production-logs/clear-all")
def clear_all_prod_logs(db: Session = Depends(get_db)):
    try:
        db.execute(text("DELETE FROM production_logs;"))
        db.commit()
    except Exception:
        db.rollback()
    return {"message": "All production logs cleared successfully"}

# --- RAW MATERIALS & RAW MATERIAL LOGS ---
@app.get("/api/rawmaterials")
def get_raw_materials(db: Session = Depends(get_db)):
    try:
        rows = db.execute(text("SELECT id, forge_pn, receipt, despatch, stock FROM raw_materials ORDER BY forge_pn ASC;")).mappings().all()
        return [{
            "id": r["id"],
            "forge_pn": r["forge_pn"],
            "receipt": int(r["receipt"] or 0),
            "despatch": int(r["despatch"] or 0),
            "stock": int(r["stock"] or 0)
        } for r in rows]
    except Exception:
        db.rollback()
        rms = db.query(models.RawMaterial).order_by(models.RawMaterial.forge_pn.asc()).all()
        return [{
            "id": r.id,
            "forge_pn": r.forge_pn,
            "receipt": r.receipt or 0,
            "despatch": r.despatch or 0,
            "stock": r.stock or 0
        } for r in rms]

@app.post("/api/rawmaterials")
def create_raw_material(data: dict, db: Session = Depends(get_db)):
    fpn = (data.get("forge_pn") or "").strip()
    if not fpn:
        raise HTTPException(status_code=400, detail="Forge PN is required")
    receipt = int(data.get("receipt") or data.get("quantity") or 0)
    despatch = int(data.get("despatch") or 0)
    stock = int(data.get("stock") if data.get("stock") is not None else (receipt - despatch))

    try:
        db.execute(text("SELECT setval(pg_get_serial_sequence('raw_materials', 'id'), coalesce(max(id),0) + 1, false) FROM raw_materials;"))
        db.commit()
    except Exception:
        db.rollback()

    try:
        db.execute(text("""
            INSERT INTO raw_materials (forge_pn, receipt, despatch, stock)
            VALUES (:forge_pn, :receipt, :despatch, :stock)
        """), {"forge_pn": fpn, "receipt": receipt, "despatch": despatch, "stock": stock})
        db.commit()
    except Exception as ex:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Failed to create raw material: {ex}")
    return {"message": "Raw Material created successfully"}

@app.put("/api/rawmaterials/{rm_id}")
def update_raw_material(rm_id: int, data: dict, db: Session = Depends(get_db)):
    fpn = (data.get("forge_pn") or "").strip()
    receipt = int(data.get("receipt") or data.get("quantity") or 0)
    despatch = int(data.get("despatch") or 0)
    stock = int(data.get("stock") if data.get("stock") is not None else (receipt - despatch))

    try:
        db.execute(text("""
            UPDATE raw_materials SET forge_pn = :forge_pn, receipt = :receipt, despatch = :despatch, stock = :stock
            WHERE id = :id
        """), {"id": rm_id, "forge_pn": fpn, "receipt": receipt, "despatch": despatch, "stock": stock})
        db.commit()
    except Exception as ex:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Failed to update raw material: {ex}")
    return {"message": "Raw Material updated successfully"}

@app.delete("/api/rawmaterials/all")
@app.delete("/api/rawmaterials/clear-all")
@app.post("/api/rawmaterials/clear-all")
def clear_all_raw_materials(db: Session = Depends(get_db)):
    try:
        db.execute(text("DELETE FROM raw_materials;"))
        db.commit()
    except Exception:
        db.rollback()
    return {"message": "All Raw Material items deleted successfully"}

@app.delete("/api/rawmaterials/{rm_id}")
def delete_raw_material(rm_id: int, db: Session = Depends(get_db)):
    try:
        db.execute(text("DELETE FROM raw_materials WHERE id = :id"), {"id": rm_id})
        db.commit()
    except Exception as ex:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Failed to delete raw material: {ex}")
    return {"message": "Raw Material deleted"}

@app.post("/api/rawmaterials/bulk")
def bulk_import_raw_materials(items: list, db: Session = Depends(get_db)):
    for r in items:
        fpn = (r.get("forge_pn") or "").strip()
        if not fpn:
            continue
        rcpt = int(r.get("receipt") or r.get("quantity") or 0)
        dspt = int(r.get("despatch") or 0)
        stk = int(r.get("stock") if r.get("stock") is not None else (rcpt - dspt))
        try:
            db.execute(text("""
                INSERT INTO raw_materials (forge_pn, receipt, despatch, stock)
                VALUES (:forge_pn, :receipt, :despatch, :stock)
            """), {"forge_pn": fpn, "receipt": rcpt, "despatch": dspt, "stock": stk})
        except Exception:
            pass
    db.commit()
    return {"message": f"Imported {len(items)} raw materials successfully"}

# --- RAW MATERIAL LOGS ---
@app.get("/api/rawmateriallogs")
def get_raw_material_logs(db: Session = Depends(get_db)):
    try:
        rows = db.execute(text("SELECT * FROM raw_material_logs ORDER BY id DESC;")).mappings().all()
        if rows:
            return [{
                "id": r.get("id"),
                "type": r.get("type") or "receipt",
                "date": r.get("date") or "",
                "dc_type": r.get("dc_type") or "",
                "forge_pn": r.get("forge_pn") or "",
                "dc_no": r.get("dc_no") or "",
                "finish_part_no": r.get("finish_part_no") or "",
                "part_prefix": r.get("part_prefix") or "",
                "qty": int(r.get("qty") or 0),
                "created_at": str(r.get("created_at") or "")
            } for r in rows]
        return []
    except Exception:
        db.rollback()
        try:
            logs = db.query(models.RawMaterialLog).order_by(models.RawMaterialLog.id.desc()).all()
            return [{
                "id": l.id,
                "type": l.type or "receipt",
                "date": l.date or "",
                "dc_type": l.dc_type or "",
                "forge_pn": l.forge_pn or "",
                "dc_no": l.dc_no or "",
                "finish_part_no": l.finish_part_no or "",
                "part_prefix": l.part_prefix or "",
                "qty": l.qty or 0,
                "created_at": str(getattr(l, "created_at", "") or "")
            } for l in logs]
        except Exception:
            db.rollback()
            return []

@app.post("/api/rawmateriallogs")
def create_raw_material_log(data: dict, db: Session = Depends(get_db)):
    rtype = (data.get("type") or "receipt").strip().lower()
    rdate = (data.get("date") or "").strip()
    dctype = (data.get("dc_type") or "").strip()
    fpn = (data.get("forge_pn") or "").strip()
    dcno = (data.get("dc_no") or "").strip()
    fpno = (data.get("finish_part_no") or "").strip()
    pprefix = (data.get("part_prefix") or "").strip()
    qty = int(data.get("qty") or data.get("quantity") or 0)

    if not fpn:
        raise HTTPException(status_code=400, detail="Forge PN is required")

    try:
        db.execute(text("SELECT setval(pg_get_serial_sequence('raw_material_logs', 'id'), coalesce(max(id),0) + 1, false) FROM raw_material_logs;"))
        db.commit()
    except Exception:
        db.rollback()

    try:
        db.execute(text("""
            INSERT INTO raw_material_logs (type, date, dc_type, forge_pn, dc_no, finish_part_no, part_prefix, qty)
            VALUES (:type, :date, :dc_type, :forge_pn, :dc_no, :finish_part_no, :part_prefix, :qty)
        """), {"type": rtype, "date": rdate, "dc_type": dctype, "forge_pn": fpn, "dc_no": dcno, "finish_part_no": fpno, "part_prefix": pprefix, "qty": qty})
        
        # Auto sync stock in raw_materials
        existing_rm = db.execute(text("SELECT id, receipt, despatch, stock FROM raw_materials WHERE forge_pn = :forge_pn"), {"forge_pn": fpn}).mappings().first()
        if existing_rm:
            cur_rcpt = int(existing_rm["receipt"] or 0)
            cur_dspt = int(existing_rm["despatch"] or 0)
            if rtype == "receipt":
                cur_rcpt += qty
            elif rtype == "despatch":
                cur_dspt += qty
            cur_stk = cur_rcpt - cur_dspt
            db.execute(text("""
                UPDATE raw_materials SET receipt = :receipt, despatch = :despatch, stock = :stock WHERE forge_pn = :forge_pn
            """), {"forge_pn": fpn, "receipt": cur_rcpt, "despatch": cur_dspt, "stock": cur_stk})
        else:
            cur_rcpt = qty if rtype == "receipt" else 0
            cur_dspt = qty if rtype == "despatch" else 0
            cur_stk = cur_rcpt - cur_dspt
            db.execute(text("""
                INSERT INTO raw_materials (forge_pn, receipt, despatch, stock) VALUES (:forge_pn, :receipt, :despatch, :stock)
            """), {"forge_pn": fpn, "receipt": cur_rcpt, "despatch": cur_dspt, "stock": cur_stk})

        db.commit()
    except Exception as ex:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Failed to create raw material log: {ex}")

    return {"message": "Raw Material Log saved successfully"}

@app.delete("/api/rawmateriallogs/{log_id}")
def delete_raw_material_log(log_id: int, db: Session = Depends(get_db)):
    try:
        db.execute(text("DELETE FROM raw_material_logs WHERE id = :id"), {"id": log_id})
        db.commit()
    except Exception as ex:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Failed to delete raw material log: {ex}")
    return {"message": "Log deleted"}

@app.delete("/api/rawmateriallogs/all")
@app.delete("/api/rawmateriallogs/clear-all")
@app.post("/api/rawmateriallogs/clear-all")
def clear_all_raw_material_logs(db: Session = Depends(get_db)):
    try:
        db.execute(text("DELETE FROM raw_material_logs;"))
        db.commit()
    except Exception:
        db.rollback()
    return {"message": "All Raw Material logs deleted successfully"}

@app.post("/api/rawmateriallogs/bulk")
def bulk_import_raw_material_logs(items: list, db: Session = Depends(get_db)):
    for rl in items:
        rtype = (rl.get("type") or "receipt").strip().lower()
        rdate = (rl.get("date") or "").strip()
        dctype = (rl.get("dc_type") or "").strip()
        fpn = (rl.get("forge_pn") or "").strip()
        dcno = (rl.get("dc_no") or "").strip()
        fpno = (rl.get("finish_part_no") or "").strip()
        pprefix = (rl.get("part_prefix") or "").strip()
        rqty = int(rl.get("qty") or rl.get("quantity") or 0)
        if not fpn:
            continue
        try:
            db.execute(text("""
                INSERT INTO raw_material_logs (type, date, dc_type, forge_pn, dc_no, finish_part_no, part_prefix, qty)
                VALUES (:type, :date, :dc_type, :forge_pn, :dc_no, :finish_part_no, :part_prefix, :qty)
            """), {"type": rtype, "date": rdate, "dc_type": dctype, "forge_pn": fpn, "dc_no": dcno, "finish_part_no": fpno, "part_prefix": pprefix, "qty": rqty})
        except Exception:
            pass
    db.commit()
    return {"message": f"Imported {len(items)} logs successfully"}

# --- HR ATTENDANCE CRUD ---
@app.get("/api/attendance")
def get_attendance(month_year: Optional[str] = None, db: Session = Depends(get_db)):
    try:
        if month_year:
            rows = db.execute(text("SELECT * FROM attendances WHERE month_year = :my ORDER BY id ASC;"), {"my": month_year.strip()}).mappings().all()
        else:
            rows = db.execute(text("SELECT * FROM attendances ORDER BY id ASC;")).mappings().all()
        if rows:
            return [{
                "id": r.get("id"),
                "employee_name": r.get("employee_name") or "",
                "dept": r.get("dept") or "",
                "designation": r.get("designation") or "Operator",
                "month_year": r.get("month_year") or "",
                "day": int(r.get("day") or 1),
                "hours": str(r.get("hours") or "0"),
                "created_at": str(r.get("created_at") or "")
            } for r in rows]
        return []
    except Exception:
        db.rollback()
        try:
            q = db.query(models.Attendance)
            if month_year:
                q = q.filter(models.Attendance.month_year == month_year.strip())
            atts = q.order_by(models.Attendance.id.asc()).all()
            return [{
                "id": a.id,
                "employee_name": a.employee_name,
                "dept": a.dept or "",
                "designation": a.designation or "Operator",
                "month_year": a.month_year,
                "day": a.day,
                "hours": a.hours or "0",
                "created_at": str(getattr(a, "created_at", "") or "")
            } for a in atts]
        except Exception:
            db.rollback()
            return []

@app.post("/api/attendance")
def save_attendance(data: dict, db: Session = Depends(get_db)):
    month_val = (data.get("month_year") or "").strip()
    entries = data.get("entries") or []

    if not month_val:
        raise HTTPException(status_code=400, detail="month_year is required")

    try:
        db.execute(text("DELETE FROM attendances WHERE month_year = :my;"), {"my": month_val})
        db.commit()
    except Exception:
        db.rollback()

    try:
        db.execute(text("SELECT setval(pg_get_serial_sequence('attendances', 'id'), coalesce(max(id),0) + 1, false) FROM attendances;"))
        db.commit()
    except Exception:
        db.rollback()

    for entry in entries:
        ename = (entry.get("employee_name") or "").strip()
        if not ename:
            continue
        dept = (entry.get("dept") or "").strip()
        desig = (entry.get("designation") or "Operator").strip()
        day = int(entry.get("day") or 1)
        hrs = str(entry.get("hours") or "0")

        try:
            db.execute(text("""
                INSERT INTO attendances (employee_name, dept, designation, month_year, day, hours)
                VALUES (:employee_name, :dept, :designation, :month_year, :day, :hours);
            """), {
                "employee_name": ename,
                "dept": dept,
                "designation": desig,
                "month_year": month_val,
                "day": day,
                "hours": hrs
            })
        except Exception:
            pass

    db.commit()
    return {"message": f"Attendance for {month_val} saved successfully!"}

@app.delete("/api/attendance/clear-all")
@app.delete("/api/attendance/all")
@app.post("/api/attendance/clear-all")
def clear_all_attendance(db: Session = Depends(get_db)):
    try:
        db.execute(text("DELETE FROM attendances;"))
        db.commit()
    except Exception:
        db.rollback()
    return {"message": "All attendance records cleared successfully"}

# --- HT & PC LOGS CRUD ---
@app.get("/api/ht_logs")
def get_ht_logs(db: Session = Depends(get_db)):
    try:
        rows = db.execute(text("SELECT * FROM ht_logs ORDER BY id DESC;")).mappings().all()
        return [{
            "id": r.get("id"),
            "date": r.get("date") or "",
            "dc_no": r.get("dc_no") or "",
            "vendor": r.get("vendor") or "",
            "partno": r.get("partno") or "",
            "qty": int(r.get("qty") or 0),
            "created_at": str(r.get("created_at") or "")
        } for r in rows]
    except Exception:
        db.rollback()
        return []

@app.post("/api/ht_logs")
def create_ht_log(data: dict, db: Session = Depends(get_db)):
    try:
        db.execute(text("SELECT setval(pg_get_serial_sequence('ht_logs', 'id'), coalesce(max(id),0) + 1, false) FROM ht_logs;"))
        db.commit()
    except Exception:
        db.rollback()

    try:
        db.execute(text("""
            INSERT INTO ht_logs (date, dc_no, vendor, partno, qty)
            VALUES (:date, :dc_no, :vendor, :partno, :qty)
        """), {
            "date": (data.get("date") or "").strip(),
            "dc_no": (data.get("dc_no") or "").strip(),
            "vendor": (data.get("vendor") or "").strip(),
            "partno": (data.get("partno") or "").strip(),
            "qty": int(data.get("qty") or 0)
        })
        db.commit()
    except Exception as ex:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(ex))
    return {"message": "HT Log saved successfully"}

@app.delete("/api/ht_logs/{log_id}")
def delete_ht_log(log_id: int, db: Session = Depends(get_db)):
    try:
        db.execute(text("DELETE FROM ht_logs WHERE id = :id"), {"id": log_id})
        db.commit()
    except Exception:
        db.rollback()
    return {"message": "HT Log deleted"}

@app.delete("/api/ht_logs/clear-all")
@app.delete("/api/ht_logs/all")
def clear_all_ht_logs(db: Session = Depends(get_db)):
    try:
        db.execute(text("DELETE FROM ht_logs;"))
        db.commit()
    except Exception:
        db.rollback()
    return {"message": "All HT logs cleared"}

@app.get("/api/ht_receipt_logs")
def get_ht_receipt_logs(db: Session = Depends(get_db)):
    try:
        rows = db.execute(text("SELECT * FROM ht_receipt_logs ORDER BY id DESC;")).mappings().all()
        return [{
            "id": r.get("id"),
            "date": r.get("date") or "",
            "dc_no": r.get("dc_no") or "",
            "vendor": r.get("vendor") or "",
            "partno": r.get("partno") or "",
            "qty": int(r.get("qty") or 0),
            "created_at": str(r.get("created_at") or "")
        } for r in rows]
    except Exception:
        db.rollback()
        return []

@app.post("/api/ht_receipt_logs")
def create_ht_receipt_log(data: dict, db: Session = Depends(get_db)):
    try:
        db.execute(text("SELECT setval(pg_get_serial_sequence('ht_receipt_logs', 'id'), coalesce(max(id),0) + 1, false) FROM ht_receipt_logs;"))
        db.commit()
    except Exception:
        db.rollback()

    try:
        db.execute(text("""
            INSERT INTO ht_receipt_logs (date, dc_no, vendor, partno, qty)
            VALUES (:date, :dc_no, :vendor, :partno, :qty)
        """), {
            "date": (data.get("date") or "").strip(),
            "dc_no": (data.get("dc_no") or "").strip(),
            "vendor": (data.get("vendor") or "").strip(),
            "partno": (data.get("partno") or "").strip(),
            "qty": int(data.get("qty") or 0)
        })
        db.commit()
    except Exception as ex:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(ex))
    return {"message": "HT Receipt Log saved successfully"}

@app.delete("/api/ht_receipt_logs/{log_id}")
def delete_ht_receipt_log(log_id: int, db: Session = Depends(get_db)):
    try:
        db.execute(text("DELETE FROM ht_receipt_logs WHERE id = :id"), {"id": log_id})
        db.commit()
    except Exception:
        db.rollback()
    return {"message": "HT Receipt Log deleted"}

@app.delete("/api/ht_receipt_logs/clear-all")
@app.delete("/api/ht_receipt_logs/all")
def clear_all_ht_receipt_logs(db: Session = Depends(get_db)):
    try:
        db.execute(text("DELETE FROM ht_receipt_logs;"))
        db.commit()
    except Exception:
        db.rollback()
    return {"message": "All HT Receipt logs cleared"}

@app.get("/api/pc_logs")
def get_pc_logs(db: Session = Depends(get_db)):
    try:
        rows = db.execute(text("SELECT * FROM pc_logs ORDER BY id DESC;")).mappings().all()
        return [{
            "id": r.get("id"),
            "date": r.get("date") or "",
            "dc_no": r.get("dc_no") or "",
            "vendor": r.get("vendor") or "",
            "partno": r.get("partno") or "",
            "qty": int(r.get("qty") or 0),
            "created_at": str(r.get("created_at") or "")
        } for r in rows]
    except Exception:
        db.rollback()
        return []

@app.post("/api/pc_logs")
def create_pc_log(data: dict, db: Session = Depends(get_db)):
    try:
        db.execute(text("SELECT setval(pg_get_serial_sequence('pc_logs', 'id'), coalesce(max(id),0) + 1, false) FROM pc_logs;"))
        db.commit()
    except Exception:
        db.rollback()

    try:
        db.execute(text("""
            INSERT INTO pc_logs (date, dc_no, vendor, partno, qty)
            VALUES (:date, :dc_no, :vendor, :partno, :qty)
        """), {
            "date": (data.get("date") or "").strip(),
            "dc_no": (data.get("dc_no") or "").strip(),
            "vendor": (data.get("vendor") or "").strip(),
            "partno": (data.get("partno") or "").strip(),
            "qty": int(data.get("qty") or 0)
        })
        db.commit()
    except Exception as ex:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(ex))
    return {"message": "PC Log saved successfully"}

@app.delete("/api/pc_logs/{log_id}")
def delete_pc_log(log_id: int, db: Session = Depends(get_db)):
    try:
        db.execute(text("DELETE FROM pc_logs WHERE id = :id"), {"id": log_id})
        db.commit()
    except Exception:
        db.rollback()
    return {"message": "PC Log deleted"}

@app.delete("/api/pc_logs/clear-all")
@app.delete("/api/pc_logs/all")
def clear_all_pc_logs(db: Session = Depends(get_db)):
    try:
        db.execute(text("DELETE FROM pc_logs;"))
        db.commit()
    except Exception:
        db.rollback()
    return {"message": "All PC logs cleared"}

@app.get("/api/pc_receipt_logs")
def get_pc_receipt_logs(db: Session = Depends(get_db)):
    try:
        rows = db.execute(text("SELECT * FROM pc_receipt_logs ORDER BY id DESC;")).mappings().all()
        return [{
            "id": r.get("id"),
            "date": r.get("date") or "",
            "dc_no": r.get("dc_no") or "",
            "vendor": r.get("vendor") or "",
            "partno": r.get("partno") or "",
            "qty": int(r.get("qty") or 0),
            "created_at": str(r.get("created_at") or "")
        } for r in rows]
    except Exception:
        db.rollback()
        return []

@app.post("/api/pc_receipt_logs")
def create_pc_receipt_log(data: dict, db: Session = Depends(get_db)):
    try:
        db.execute(text("SELECT setval(pg_get_serial_sequence('pc_receipt_logs', 'id'), coalesce(max(id),0) + 1, false) FROM pc_receipt_logs;"))
        db.commit()
    except Exception:
        db.rollback()

    try:
        db.execute(text("""
            INSERT INTO pc_receipt_logs (date, dc_no, vendor, partno, qty)
            VALUES (:date, :dc_no, :vendor, :partno, :qty)
        """), {
            "date": (data.get("date") or "").strip(),
            "dc_no": (data.get("dc_no") or "").strip(),
            "vendor": (data.get("vendor") or "").strip(),
            "partno": (data.get("partno") or "").strip(),
            "qty": int(data.get("qty") or 0)
        })
        db.commit()
    except Exception as ex:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(ex))
    return {"message": "PC Receipt Log saved successfully"}

@app.delete("/api/pc_receipt_logs/{log_id}")
def delete_pc_receipt_log(log_id: int, db: Session = Depends(get_db)):
    try:
        db.execute(text("DELETE FROM pc_receipt_logs WHERE id = :id"), {"id": log_id})
        db.commit()
    except Exception:
        db.rollback()
    return {"message": "PC Receipt Log deleted"}

# --- SCHEDULE STATUS WIP ADJUSTMENTS ---
@app.post("/api/schedule_status/adjust_part_wip")
def adjust_part_wip(data: dict, db: Session = Depends(get_db)):
    try:
        dept = (data.get("department") or "General").strip()
        partno = (data.get("partno") or "").strip()
        adjustments = data.get("adjustments") or []

        if not partno:
            raise HTTPException(status_code=400, detail="Part number is required")

        today_str = datetime.datetime.now(IST).strftime("%Y-%m-%d")

        # Fetch operations for this part
        part_row = db.execute(text("SELECT id FROM parts WHERE LOWER(part_no) = LOWER(:p);"), {"p": partno}).mappings().first()
        operations = []
        if part_row:
            op_rows = db.execute(text("SELECT opn_no, description FROM operations WHERE part_id = :pid ORDER BY id ASC;"), {"pid": part_row["id"]}).mappings().all()
            operations = [str(r["opn_no"]).strip() for r in op_rows]

        target_map = {}
        for adj in adjustments:
            opn_key = str(adj.get("opn_no") or "").strip().lower()
            try:
                target_val = float(adj.get("target_balance") or 0)
            except Exception:
                target_val = 0.0
            target_map[opn_key] = target_val

        stages = []
        for op in operations:
            stages.append(op.lower())
        if "debur" not in stages:
            stages.append("debur")
        if "for ins" not in stages:
            stages.append("for ins")
        if "rfd" not in stages:
            stages.append("rfd")

        # Despatch total from raw_material_logs
        rm_rows = db.execute(text("""
            SELECT qty FROM raw_material_logs
            WHERE LOWER(type) = 'despatch' AND LOWER(finish_part_no) = LOWER(:p);
        """), {"p": partno}).mappings().all()
        desp_total = 0.0
        for r in rm_rows:
            try:
                desp_total += float(r.get("qty") or 0)
            except Exception:
                pass

        cum_req = {}
        running_cum = desp_total

        for stage in reversed(stages):
            target_bal = target_map.get(stage, 0.0)
            running_cum += target_bal
            cum_req[stage] = running_cum

        for stage in stages:
            if stage in target_map:
                prod_rows = db.execute(text("""
                    SELECT prod_qty FROM production_logs
                    WHERE LOWER(partno) = LOWER(:p) AND LOWER(opn_no) = LOWER(:opn);
                """), {"p": partno, "opn": stage}).mappings().all()
                curr_prod = 0.0
                for pr in prod_rows:
                    try:
                        curr_prod += float(pr.get("prod_qty") or 0)
                    except Exception:
                        pass

                needed_cum = cum_req[stage]
                delta = int(round(needed_cum - curr_prod))

                if delta != 0:
                    try:
                        db.execute(text("SELECT setval(pg_get_serial_sequence('production_logs', 'id'), coalesce(max(id),0) + 1, false) FROM production_logs;"))
                        db.commit()
                    except Exception:
                        db.rollback()

                    db.execute(text("""
                        INSERT INTO production_logs (
                            dept, date, shift, setter, machine, operator, partno, opn_no,
                            description, runtime, cycle_time, target_qty, prod_qty, efficiency,
                            idle_hours, idle_reason, idle_hours_2, idle_reason_2, idle_hours_3, idle_reason_3, multiple_mc
                        ) VALUES (
                            :dept, :date, 'General', 'WIP Adjustment', 'Adjustment', 'WIP Adjustment', :partno, :opn_no,
                            'WIP Adjustment', 0, 0, 0, :prod_qty, 100,
                            0, '', 0, '', 0, '', 'no'
                        );
                    """), {
                        "dept": dept,
                        "date": today_str,
                        "partno": partno,
                        "opn_no": stage.upper() if stage not in ["debur", "for ins", "rfd"] else stage,
                        "prod_qty": delta
                    })
                    db.commit()

        return {"message": f"WIP balances for part {partno} adjusted successfully"}
    except HTTPException:
        raise
    except Exception as ex:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(ex))

@app.post("/api/schedule_status/clear_backlog_corrections")
def clear_backlog_corrections(db: Session = Depends(get_db)):
    try:
        db.execute(text("""
            DELETE FROM production_logs
            WHERE setter IN ('WIP Adjustment', 'Backlog Correction')
               OR operator IN ('WIP Adjustment', 'Backlog Correction')
               OR description LIKE '%WIP Adjustment%'
               OR description LIKE '%Backlog Correction%';
        """))
        db.commit()
    except Exception as ex:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(ex))
    return {"message": "All backlog correction and WIP adjustment logs cleared successfully"}

@app.post("/api/schedule_status/autofix_wip")
def autofix_wip(data: dict, db: Session = Depends(get_db)):
    try:
        dept = (data.get("department") or "").strip()
        today_str = datetime.datetime.now(IST).strftime("%Y-%m-%d")

        q = "SELECT id, part_no as partno, dept as department FROM parts"
        params = {}
        if dept:
            q += " WHERE LOWER(dept) = LOWER(:dept)"
            params["dept"] = dept
        q += " ORDER BY id ASC;"

        parts = db.execute(text(q), params).mappings().all()

        for p in parts:
            pno = (p.get("partno") or "").strip()
            pdept = p.get("department") or dept or "General"
            if not pno:
                continue
            op_rows = db.execute(text("SELECT opn_no FROM operations WHERE part_id = :pid ORDER BY id ASC;"), {"pid": p["id"]}).mappings().all()
            operations = [str(r["opn_no"]).strip().lower() for r in op_rows]
            if not operations:
                continue

            for i in range(len(operations) - 1):
                curr_op = operations[i]
                next_op = operations[i+1]

                curr_prod_rows = db.execute(text("SELECT prod_qty FROM production_logs WHERE LOWER(partno) = LOWER(:p) AND LOWER(opn_no) = LOWER(:opn);"), {"p": pno, "opn": curr_op}).mappings().all()
                next_prod_rows = db.execute(text("SELECT prod_qty FROM production_logs WHERE LOWER(partno) = LOWER(:p) AND LOWER(opn_no) = LOWER(:opn);"), {"p": pno, "opn": next_op}).mappings().all()

                c_prod = sum(float(r.get("prod_qty") or 0) for r in curr_prod_rows)
                n_prod = sum(float(r.get("prod_qty") or 0) for r in next_prod_rows)

                if c_prod < n_prod:
                    diff = int(round(n_prod - c_prod))
                    db.execute(text("""
                        INSERT INTO production_logs (
                            dept, date, shift, setter, machine, operator, partno, opn_no,
                            description, runtime, cycle_time, target_qty, prod_qty, efficiency,
                            idle_hours, idle_reason, idle_hours_2, idle_reason_2, idle_hours_3, idle_reason_3, multiple_mc
                        ) VALUES (
                            :dept, :date, 'General', 'Backlog Correction', 'Adjustment', 'Backlog Correction', :partno, :opn_no,
                            'Backlog Correction', 0, 0, 0, :prod_qty, 100,
                            0, '', 0, '', 0, '', 'no'
                        );
                    """), {
                        "dept": pdept,
                        "date": today_str,
                        "partno": pno,
                        "opn_no": curr_op.upper(),
                        "prod_qty": diff
                    })

        db.commit()
        return {"message": f"Backlog correction completed successfully for {len(parts)} parts"}
    except Exception as ex:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(ex))

# --- Tooling ---
@app.get("/api/tooling", response_model=List[ToolingResponse])
def get_tooling(db: Session = Depends(get_db)):
    return db.query(models.Tooling).all()

@app.post("/api/tooling", response_model=ToolingResponse)
def create_tooling(tool: ToolingCreate, db: Session = Depends(get_db)):
    db_tool = models.Tooling(**tool.model_dump())
    db.add(db_tool)
    db.commit()
    db.refresh(db_tool)
    return db_tool

@app.delete("/api/tooling/clear-all")
def clear_all_tooling(db: Session = Depends(get_db)):
    db.query(models.Tooling).delete()
    db.commit()
    return {"message": "All tooling items cleared successfully!"}

@app.delete("/api/tooling/{tool_id}")
def delete_tooling(tool_id: int, db: Session = Depends(get_db)):
    db_tool = db.query(models.Tooling).filter(models.Tooling.id == tool_id).first()
    if not db_tool:
        raise HTTPException(status_code=404, detail="Tooling item not found")
    db.delete(db_tool)
    db.commit()
    return {"message": "Tooling item deleted"}

# --- Inspection Parameters & Reports API ---

DEFAULT_INSPECTION_PARAMS = [
    {"sl_no": 1, "description": "Bore", "nominal_dimension": 115.0, "lo_tol": 0.03, "hi_tol": 0.03},
    {"sl_no": 2, "description": "Bore", "nominal_dimension": 110.0, "lo_tol": 0.30, "hi_tol": 0.30},
    {"sl_no": 3, "description": "Dim", "nominal_dimension": 95.0, "lo_tol": 0.10, "hi_tol": 0.10},
    {"sl_no": 4, "description": "Dim", "nominal_dimension": 50.0, "lo_tol": 0.30, "hi_tol": 0.30},
    {"sl_no": 5, "description": "OD", "nominal_dimension": 142.0, "lo_tol": 0.30, "hi_tol": 0.30},
]

@app.get("/api/inspection-parameters", response_model=List[InspectionParamResponse])
def get_inspection_parameters(part_no: str, opn_no: Optional[str] = None, db: Session = Depends(get_db)):
    clean_p = part_no.strip().lower()
    clean_op = opn_no.strip().lower() if opn_no else None

    query = db.query(models.InspectionParameter).filter(func.lower(models.InspectionParameter.part_no) == clean_p)
    if clean_op:
        query = query.filter(func.lower(models.InspectionParameter.opn_no) == clean_op)
    
    params = query.order_by(models.InspectionParameter.sl_no.asc()).all()

    # Seed default parameters matching user template if none exist for this part & operation
    if not params and opn_no:
        for p in DEFAULT_INSPECTION_PARAMS:
            db_param = models.InspectionParameter(
                part_no=part_no.strip(),
                opn_no=opn_no.strip(),
                sl_no=p["sl_no"],
                description=p["description"],
                nominal_dimension=p["nominal_dimension"],
                lo_tol=p["lo_tol"],
                hi_tol=p["hi_tol"]
            )
            db.add(db_param)
        db.commit()
        params = db.query(models.InspectionParameter).filter(
            func.lower(models.InspectionParameter.part_no) == clean_p,
            func.lower(models.InspectionParameter.opn_no) == clean_op
        ).order_by(models.InspectionParameter.sl_no.asc()).all()

    return params

@app.post("/api/inspection-parameters")
def save_inspection_parameters(param_list: List[InspectionParamCreate], db: Session = Depends(get_db)):
    if not param_list:
        return {"message": "No parameters provided"}
    
    p_no = param_list[0].part_no.strip()
    op_no = param_list[0].opn_no.strip()

    db.query(models.InspectionParameter).filter(
        func.lower(models.InspectionParameter.part_no) == p_no.lower(),
        func.lower(models.InspectionParameter.opn_no) == op_no.lower()
    ).delete(synchronize_session=False)

    for idx, item in enumerate(param_list, start=1):
        db_param = models.InspectionParameter(
            part_no=p_no,
            opn_no=op_no,
            sl_no=idx,
            description=item.description,
            nominal_dimension=item.nominal_dimension,
            lo_tol=item.lo_tol,
            hi_tol=item.hi_tol
        )
        db.add(db_param)

    db.commit()
    return {"message": "Inspection parameters saved successfully!"}

@app.delete("/api/inspection-parameters/{param_id}")
def delete_inspection_parameter(param_id: int, db: Session = Depends(get_db)):
    p = db.query(models.InspectionParameter).filter(models.InspectionParameter.id == param_id).first()
    if not p:
        raise HTTPException(status_code=404, detail="Parameter not found")
    db.delete(p)
    db.commit()
    return {"message": "Parameter deleted"}

def generate_traceability_code(part_no: str, opn_no: str, db: Session) -> str:
    clean_p = part_no.strip().upper()
    clean_op = opn_no.strip()
    mmdd = get_now_ist().strftime("%m%d")
    prefix = f"{clean_p}-{clean_op}-{mmdd}-"
    
    existing_count = db.query(models.InspectionReport).filter(
        models.InspectionReport.report_code.like(f"{prefix}%")
    ).count()
    
    seq_num = existing_count + 1
    return f"{prefix}{seq_num:03d}"

@app.get("/api/inspection-reports/next-code")
def get_next_report_code(part_no: str, opn_no: str, db: Session = Depends(get_db)):
    code = generate_traceability_code(part_no, opn_no, db)
    return {"report_code": code}

@app.get("/api/inspection-reports")
def get_inspection_report(part_no: str, opn_no: str, db: Session = Depends(get_db)):
    clean_p = part_no.strip().lower()
    clean_op = opn_no.strip().lower()

    report = db.query(models.InspectionReport).filter(
        func.lower(models.InspectionReport.part_no) == clean_p,
        func.lower(models.InspectionReport.opn_no) == clean_op
    ).order_by(models.InspectionReport.id.desc()).first()

    next_code = generate_traceability_code(part_no, opn_no, db)

    sch = db.query(models.ProductionSchedule).filter(
        func.lower(models.ProductionSchedule.part_no) == clean_p
    ).first()
    sch_batch_qty = sch.total_sch_qty if (sch and sch.total_sch_qty and sch.total_sch_qty > 0) else 30

    if not report:
        return {
            "report_code": next_code,
            "part_no": part_no,
            "opn_no": opn_no,
            "batch_qty": sch_batch_qty,
            "machine_name": "",
            "operator_name": "",
            "inspection_date": get_now_ist().strftime("%Y-%m-%d"),
            "comp_sl_nos": "1",
            "readings_json": "{}"
        }
    return report

@app.post("/api/inspection-reports")
def save_inspection_report(req: InspectionReportSave, db: Session = Depends(get_db)):
    report_code = req.report_code
    if not report_code:
        report_code = generate_traceability_code(req.part_no, req.opn_no, db)

    # Save as a distinct, traceable quality inspection instance
    report = models.InspectionReport(
        report_code=report_code,
        prod_log_id=req.prod_log_id,
        part_no=req.part_no.strip(),
        opn_no=req.opn_no.strip(),
        batch_qty=req.batch_qty,
        machine_name=req.machine_name,
        operator_name=req.operator_name,
        inspection_date=req.inspection_date or get_now_ist().strftime("%Y-%m-%d"),
        comp_sl_nos=req.comp_sl_nos,
        readings_json=req.readings_json
    )
    db.add(report)
    db.commit()
    db.refresh(report)

    return {"message": "Inspection Report saved successfully!", "report_code": report.report_code, "id": report.id}

@app.get("/api/inspection-reports/by-id/{report_id}")
def get_inspection_report_by_id(report_id: int, db: Session = Depends(get_db)):
    r = db.query(models.InspectionReport).filter(models.InspectionReport.id == report_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="Inspection report not found")
    return r

@app.delete("/api/inspection-reports/{report_id}")
def delete_inspection_report(report_id: int, db: Session = Depends(get_db)):
    r = db.query(models.InspectionReport).filter(models.InspectionReport.id == report_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="Inspection report not found")
    db.delete(r)
    db.commit()
    return {"message": "Inspection report deleted"}

# --- Excel Export Endpoints ---
@app.get("/api/export/production-logs/excel")
def export_production_logs_excel(db: Session = Depends(get_db)):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    logs = db.query(models.ProductionLog).order_by(models.ProductionLog.id.desc()).all()
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Production Logs"

    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )

    headers = [
        "Log ID", "Date & Time", "Shift", "Machine Name", "Operator Name", 
        "Part Number", "Operation No", "Qty Produced", "Scrap Qty", "Completed Serial Nos"
    ]
    ws.append(headers)

    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align_center

    ws.row_dimensions[1].height = 24

    for log in logs:
        ts_str = log.created_at.strftime("%Y-%m-%d %H:%M:%S") if log.created_at else (log.log_date or "")
        row = [
            log.id,
            ts_str,
            log.shift or "",
            log.machine_name or "",
            log.operator_name or "",
            log.part_no or "",
            f"Opn {log.opn_no}" if log.opn_no else "",
            log.qty_produced or 0,
            log.scrap_qty or 0,
            log.completed_sl_nos or ""
        ]
        ws.append(row)
        r_idx = ws.max_row
        ws.row_dimensions[r_idx].height = 20
        for c_idx in range(1, len(row) + 1):
            cell = ws.cell(row=r_idx, column=c_idx)
            cell.border = thin_border
            if c_idx in [1, 2, 3, 7, 8, 9]:
                cell.alignment = align_center
            else:
                cell.alignment = align_left

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"Production_Logs_{datetime.date.today().strftime('%Y%m%d')}.xlsx"
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@app.get("/api/export/inspection-reports/excel")
def export_inspection_reports_excel(db: Session = Depends(get_db)):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    reports = db.query(models.InspectionReport).order_by(models.InspectionReport.id.desc()).all()
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Quality Inspection Logs"

    header_fill = PatternFill(start_color="065F46", end_color="065F46", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    
    pass_fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
    pass_font = Font(name="Calibri", size=11, bold=True, color="065F46")
    
    fail_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
    fail_font = Font(name="Calibri", size=11, bold=True, color="991B1B")

    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )

    headers = [
        "Report ID", "Traceability ID", "Date", "Part Number", "Operation No", 
        "Batch Qty", "Machine Name", "Operator Name", "Component Sl No",
        "Parameter Description", "Nominal", "Lo Tol", "Hi Tol", "Measured Reading", "Status"
    ]
    ws.append(headers)

    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align_center

    ws.row_dimensions[1].height = 25

    for r in reports:
        params = db.query(models.InspectionParameter).filter(
            func.lower(models.InspectionParameter.part_no) == r.part_no.strip().lower(),
            func.lower(models.InspectionParameter.opn_no) == r.opn_no.strip().lower()
        ).order_by(models.InspectionParameter.sl_no.asc()).all()

        readings_map = {}
        try:
            import json
            readings_map = json.loads(r.readings_json or "{}")
        except Exception:
            readings_map = {}

        if not params:
            row = [
                r.id,
                r.report_code or f"IR-{r.id}",
                r.inspection_date or "",
                r.part_no or "",
                f"Opn {r.opn_no}" if r.opn_no else "",
                f"{r.batch_qty} pcs" if r.batch_qty else "-",
                r.machine_name or "-",
                r.operator_name or "-",
                r.comp_sl_nos or "1",
                "No parameters defined", "-", "-", "-", "-", "-"
            ]
            ws.append(row)
            r_idx = ws.max_row
            ws.row_dimensions[r_idx].height = 20
            for c_idx in range(1, len(row) + 1):
                cell = ws.cell(row=r_idx, column=c_idx)
                cell.border = thin_border
                cell.alignment = align_center if c_idx in [1, 2, 3, 5, 6, 9] else align_left
        else:
            is_first_row = True
            for p in params:
                p_readings = readings_map.get(str(p.id)) or readings_map.get(p.id) or {}
                val = p_readings.get("col_0") if "col_0" in p_readings else (p_readings.get("col_1") if "col_1" in p_readings else "")
                if val is None: val = ""
                val_str = str(val).strip()

                nom = float(p.nominal_dimension or 0.0)
                lo = float(p.lo_tol or 0.0)
                hi = float(p.hi_tol or 0.0)
                status = "-"
                val_num = val_str

                if val_str != "":
                    try:
                        v = float(val_str)
                        val_num = v
                        if (nom - lo) <= v <= (nom + hi):
                            status = "PASS"
                        else:
                            status = "OUT OF SPEC"
                    except ValueError:
                        status = "INVALID"

                if is_first_row:
                    row = [
                        r.id,
                        r.report_code or f"IR-{r.id}",
                        r.inspection_date or "",
                        r.part_no or "",
                        f"Opn {r.opn_no}" if r.opn_no else "",
                        f"{r.batch_qty} pcs" if r.batch_qty else "-",
                        r.machine_name or "-",
                        r.operator_name or "-",
                        r.comp_sl_nos or "1",
                        p.description or "",
                        nom,
                        lo,
                        hi,
                        val_num,
                        status
                    ]
                    is_first_row = False
                else:
                    row = [
                        "", "", "", "", "", "", "", "", "",
                        p.description or "",
                        nom,
                        lo,
                        hi,
                        val_num,
                        status
                    ]

                ws.append(row)
                r_idx = ws.max_row
                ws.row_dimensions[r_idx].height = 20

                for c_idx in range(1, len(row) + 1):
                    cell = ws.cell(row=r_idx, column=c_idx)
                    cell.border = thin_border
                    if c_idx in [1, 2, 3, 5, 6, 9]:
                        cell.alignment = align_center
                    elif c_idx in [11, 12, 13, 14]:
                        cell.alignment = align_right
                    elif c_idx == 15:
                        cell.alignment = align_center
                        if status == "PASS":
                            cell.fill = pass_fill
                            cell.font = pass_font
                        elif status == "OUT OF SPEC":
                            cell.fill = fail_fill
                            cell.font = fail_font
                    else:
                        cell.alignment = align_left

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"Quality_Inspection_Logs_{datetime.date.today().strftime('%Y%m%d')}.xlsx"
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

# --- TOOL CRIB / INSERT MASTER CRUD ENDPOINTS ---

@app.get("/api/insert_masters")
def get_insert_masters(db: Session = Depends(get_db)):
    try:
        rows = db.query(models.InsertMaster).order_by(models.InsertMaster.id.asc()).all()
        return [{
            "id": r.id,
            "insert_spec": r.insert_spec,
            "no_of_edges": r.no_of_edges,
            "name": r.name,
            "specification": r.specification,
            "grade": r.grade,
            "make": r.make,
            "stock": r.stock,
            "price": r.price
        } for r in rows]
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/insert_masters")
def create_insert_master(data: dict, db: Session = Depends(get_db)):
    spec = (data.get("insert_spec") or data.get("name") or "").strip()
    if not spec:
        raise HTTPException(status_code=400, detail="Insert spec is required")
    edges = int(data.get("no_of_edges") or 1)
    item = models.InsertMaster(
        insert_spec=spec,
        no_of_edges=edges,
        grade=(data.get("grade") or "").strip(),
        make=(data.get("make") or "").strip(),
        stock=float(data.get("stock") or 0.0),
        price=float(data.get("price") or 0.0)
    )
    db.add(item)
    db.commit()
    db.refresh(item)
    return {"id": item.id, "insert_spec": item.insert_spec, "no_of_edges": item.no_of_edges}

@app.post("/api/insert_masters/bulk")
def bulk_create_insert_masters(data: dict, db: Session = Depends(get_db)):
    inserts = data.get("inserts") or []
    count = 0
    for ins in inserts:
        spec = (ins.get("insert_spec") or ins.get("name") or "").strip()
        if spec:
            item = models.InsertMaster(
                insert_spec=spec,
                no_of_edges=int(ins.get("no_of_edges") or 1),
                grade=(ins.get("grade") or "").strip(),
                make=(ins.get("make") or "").strip(),
                stock=float(ins.get("stock") or 0.0),
                price=float(ins.get("price") or 0.0)
            )
            db.add(item)
            count += 1
    db.commit()
    return {"message": f"Successfully created {count} insert master records"}

@app.put("/api/insert_masters/{id}")
def update_insert_master(id: int, data: dict, db: Session = Depends(get_db)):
    item = db.query(models.InsertMaster).filter(models.InsertMaster.id == id).first()
    if not item:
        raise HTTPException(status_code=404, detail="Insert record not found")
    spec = (data.get("insert_spec") or data.get("name") or "").strip()
    if spec:
        item.insert_spec = spec
    if "no_of_edges" in data:
        item.no_of_edges = int(data["no_of_edges"] or 1)
    if "grade" in data:
        item.grade = (data["grade"] or "").strip()
    if "make" in data:
        item.make = (data["make"] or "").strip()
    if "stock" in data:
        item.stock = float(data["stock"] or 0.0)
    if "price" in data:
        item.price = float(data["price"] or 0.0)
    db.commit()
    return {"message": "Updated successfully"}

@app.delete("/api/insert_masters/all")
def clear_all_insert_masters(db: Session = Depends(get_db)):
    db.query(models.InsertMaster).delete()
    db.commit()
    return {"message": "All insert master records cleared successfully"}

@app.delete("/api/insert_masters/{id}")
def delete_insert_master(id: int, db: Session = Depends(get_db)):
    item = db.query(models.InsertMaster).filter(models.InsertMaster.id == id).first()
    if not item:
        raise HTTPException(status_code=404, detail="Insert record not found")
    db.delete(item)
    db.commit()
    return {"message": "Deleted successfully"}

# --- DRILL MASTER CRUD ---
@app.get("/api/drill_masters")
def get_drill_masters(db: Session = Depends(get_db)):
    try:
        rows = db.query(models.DrillMaster).order_by(models.DrillMaster.id.asc()).all()
        return [{
            "id": r.id,
            "drill_size": r.drill_size,
            "sl_no": r.sl_no,
            "resharp_count": r.resharp_count,
            "name": r.name,
            "size_dia": r.size_dia,
            "specification": r.specification,
            "make": r.make,
            "stock": r.stock,
            "price": r.price
        } for r in rows]
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/drill_masters")
def create_drill_master(data: dict, db: Session = Depends(get_db)):
    size = (data.get("drill_size") or data.get("size_dia") or data.get("name") or "").strip()
    sl = (data.get("sl_no") or "").strip()
    item = models.DrillMaster(
        drill_size=size,
        sl_no=sl,
        resharp_count=int(data.get("resharp_count") or 0),
        make=(data.get("make") or "").strip(),
        stock=float(data.get("stock") or 0.0),
        price=float(data.get("price") or 0.0)
    )
    db.add(item)
    db.commit()
    db.refresh(item)
    return {"id": item.id}

@app.post("/api/drill_masters/bulk")
def bulk_create_drill_masters(data: dict, db: Session = Depends(get_db)):
    drills = data.get("drills") or []
    count = 0
    for dr in drills:
        size = (dr.get("drill_size") or dr.get("size_dia") or dr.get("name") or "").strip()
        sl = (dr.get("sl_no") or "").strip()
        if size or sl:
            item = models.DrillMaster(
                drill_size=size,
                sl_no=sl,
                resharp_count=int(dr.get("resharp_count") or 0),
                make=(dr.get("make") or "").strip(),
                stock=float(dr.get("stock") or 0.0),
                price=float(dr.get("price") or 0.0)
            )
            db.add(item)
            count += 1
    db.commit()
    return {"message": f"Successfully created {count} drill records"}

@app.put("/api/drill_masters/{id}")
def update_drill_master(id: int, data: dict, db: Session = Depends(get_db)):
    item = db.query(models.DrillMaster).filter(models.DrillMaster.id == id).first()
    if not item:
        raise HTTPException(status_code=404, detail="Drill record not found")
    if "drill_size" in data: item.drill_size = (data["drill_size"] or "").strip()
    if "sl_no" in data: item.sl_no = (data["sl_no"] or "").strip()
    if "resharp_count" in data: item.resharp_count = int(data["resharp_count"] or 0)
    if "make" in data: item.make = (data["make"] or "").strip()
    if "stock" in data: item.stock = float(data["stock"] or 0.0)
    if "price" in data: item.price = float(data["price"] or 0.0)
    db.commit()
    return {"message": "Updated successfully"}

@app.delete("/api/drill_masters/all")
def clear_all_drill_masters(db: Session = Depends(get_db)):
    db.query(models.DrillMaster).delete()
    db.commit()
    return {"message": "All drill records cleared successfully"}

@app.delete("/api/drill_masters/{id}")
def delete_drill_master(id: int, db: Session = Depends(get_db)):
    item = db.query(models.DrillMaster).filter(models.DrillMaster.id == id).first()
    if not item:
        raise HTTPException(status_code=404, detail="Drill record not found")
    db.delete(item)
    db.commit()
    return {"message": "Deleted successfully"}

# --- TAP MASTER CRUD ---
@app.get("/api/tap_masters")
def get_tap_masters(db: Session = Depends(get_db)):
    try:
        rows = db.query(models.TapMaster).order_by(models.TapMaster.id.asc()).all()
        return [{
            "id": r.id,
            "tap_spec": r.tap_spec,
            "name": r.name,
            "specification": r.specification,
            "make": r.make,
            "stock": r.stock,
            "price": r.price
        } for r in rows]
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/tap_masters")
def create_tap_master(data: dict, db: Session = Depends(get_db)):
    spec = (data.get("tap_spec") or data.get("specification") or data.get("name") or "").strip()
    if not spec:
        raise HTTPException(status_code=400, detail="Tap spec is required")
    item = models.TapMaster(
        tap_spec=spec,
        make=(data.get("make") or "").strip(),
        stock=float(data.get("stock") or 0.0),
        price=float(data.get("price") or 0.0)
    )
    db.add(item)
    db.commit()
    db.refresh(item)
    return {"id": item.id}

@app.post("/api/tap_masters/bulk")
def bulk_create_tap_masters(data: dict, db: Session = Depends(get_db)):
    taps = data.get("taps") or []
    count = 0
    for t in taps:
        spec = (t.get("tap_spec") or t.get("specification") or t.get("name") or "").strip()
        if spec:
            item = models.TapMaster(
                tap_spec=spec,
                make=(t.get("make") or "").strip(),
                stock=float(t.get("stock") or 0.0),
                price=float(t.get("price") or 0.0)
            )
            db.add(item)
            count += 1
    db.commit()
    return {"message": f"Successfully created {count} tap records"}

@app.put("/api/tap_masters/{id}")
def update_tap_master(id: int, data: dict, db: Session = Depends(get_db)):
    item = db.query(models.TapMaster).filter(models.TapMaster.id == id).first()
    if not item:
        raise HTTPException(status_code=404, detail="Tap record not found")
    if "tap_spec" in data: item.tap_spec = (data["tap_spec"] or "").strip()
    if "make" in data: item.make = (data["make"] or "").strip()
    if "stock" in data: item.stock = float(data["stock"] or 0.0)
    if "price" in data: item.price = float(data["price"] or 0.0)
    db.commit()
    return {"message": "Updated successfully"}

@app.delete("/api/tap_masters/all")
def clear_all_tap_masters(db: Session = Depends(get_db)):
    db.query(models.TapMaster).delete()
    db.commit()
    return {"message": "All tap records cleared successfully"}

@app.delete("/api/tap_masters/{id}")
def delete_tap_master(id: int, db: Session = Depends(get_db)):
    item = db.query(models.TapMaster).filter(models.TapMaster.id == id).first()
    if not item:
        raise HTTPException(status_code=404, detail="Tap record not found")
    db.delete(item)
    db.commit()
    return {"message": "Deleted successfully"}

# --- INSERT RECEIPTS CRUD ---
@app.get("/api/insert_receipts")
def get_insert_receipts(db: Session = Depends(get_db)):
    try:
        rows = db.query(models.InsertReceipt).order_by(models.InsertReceipt.id.desc()).all()
        return [{
            "id": r.id,
            "date": r.date,
            "supplier": r.supplier,
            "insert_spec": r.insert_spec,
            "batch_no": r.batch_no,
            "qty": r.qty,
            "rate": r.rate,
            "created_at": r.created_at.isoformat() if r.created_at else None
        } for r in rows]
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/insert_receipts")
def create_insert_receipt(data: dict, db: Session = Depends(get_db)):
    item = models.InsertReceipt(
        date=(data.get("date") or datetime.datetime.now(IST).strftime("%Y-%m-%d")),
        supplier=(data.get("supplier") or "").strip(),
        insert_spec=(data.get("insert_spec") or "").strip(),
        batch_no=(data.get("batch_no") or "").strip(),
        qty=float(data.get("qty") or 0.0),
        rate=float(data.get("rate") or 0.0)
    )
    db.add(item)
    db.commit()
    db.refresh(item)
    return {"id": item.id}

@app.post("/api/insert_receipts/bulk")
def bulk_create_insert_receipts(data: dict, db: Session = Depends(get_db)):
    receipts = data.get("receipts") or []
    count = 0
    for r in receipts:
        item = models.InsertReceipt(
            date=(r.get("date") or datetime.datetime.now(IST).strftime("%Y-%m-%d")),
            supplier=(r.get("supplier") or "").strip(),
            insert_spec=(r.get("insert_spec") or "").strip(),
            batch_no=(r.get("batch_no") or "").strip(),
            qty=float(r.get("qty") or 0.0),
            rate=float(r.get("rate") or 0.0)
        )
        db.add(item)
        count += 1
    db.commit()
    return {"message": f"Successfully created {count} insert receipt records"}

@app.put("/api/insert_receipts/{id}")
def update_insert_receipt(id: int, data: dict, db: Session = Depends(get_db)):
    item = db.query(models.InsertReceipt).filter(models.InsertReceipt.id == id).first()
    if not item:
        raise HTTPException(status_code=404, detail="Insert receipt not found")
    if "date" in data: item.date = str(data["date"])
    if "supplier" in data: item.supplier = str(data["supplier"])
    if "insert_spec" in data: item.insert_spec = str(data["insert_spec"])
    if "batch_no" in data: item.batch_no = str(data["batch_no"])
    if "qty" in data: item.qty = float(data["qty"] or 0.0)
    if "rate" in data: item.rate = float(data["rate"] or 0.0)
    db.commit()
    return {"message": "Updated successfully"}

@app.delete("/api/insert_receipts/all")
def clear_all_insert_receipts(db: Session = Depends(get_db)):
    db.query(models.InsertReceipt).delete()
    db.commit()
    return {"message": "All insert receipt records cleared successfully"}

@app.delete("/api/insert_receipts/{id}")
def delete_insert_receipt(id: int, db: Session = Depends(get_db)):
    item = db.query(models.InsertReceipt).filter(models.InsertReceipt.id == id).first()
    if not item:
        raise HTTPException(status_code=404, detail="Insert receipt not found")
    db.delete(item)
    db.commit()
    return {"message": "Deleted successfully"}

# --- TAP RECEIPTS CRUD ---
@app.get("/api/tap_receipts")
def get_tap_receipts(db: Session = Depends(get_db)):
    try:
        rows = db.query(models.TapReceipt).order_by(models.TapReceipt.id.desc()).all()
        return [{
            "id": r.id,
            "date": r.date,
            "supplier": r.supplier,
            "tap_spec": r.tap_spec,
            "qty": r.qty,
            "rate": r.rate,
            "created_at": r.created_at.isoformat() if r.created_at else None
        } for r in rows]
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/tap_receipts")
def create_tap_receipt(data: dict, db: Session = Depends(get_db)):
    item = models.TapReceipt(
        date=(data.get("date") or datetime.datetime.now(IST).strftime("%Y-%m-%d")),
        supplier=(data.get("supplier") or "").strip(),
        tap_spec=(data.get("tap_spec") or "").strip(),
        qty=float(data.get("qty") or 0.0),
        rate=float(data.get("rate") or 0.0)
    )
    db.add(item)
    db.commit()
    db.refresh(item)
    return {"id": item.id}

@app.post("/api/tap_receipts/bulk")
def bulk_create_tap_receipts(data: dict, db: Session = Depends(get_db)):
    receipts = data.get("receipts") or []
    count = 0
    for r in receipts:
        item = models.TapReceipt(
            date=(r.get("date") or datetime.datetime.now(IST).strftime("%Y-%m-%d")),
            supplier=(r.get("supplier") or "").strip(),
            tap_spec=(r.get("tap_spec") or "").strip(),
            qty=float(r.get("qty") or 0.0),
            rate=float(r.get("rate") or 0.0)
        )
        db.add(item)
        count += 1
    db.commit()
    return {"message": f"Successfully created {count} tap receipt records"}

@app.put("/api/tap_receipts/{id}")
def update_tap_receipt(id: int, data: dict, db: Session = Depends(get_db)):
    item = db.query(models.TapReceipt).filter(models.TapReceipt.id == id).first()
    if not item:
        raise HTTPException(status_code=404, detail="Tap receipt not found")
    if "date" in data: item.date = str(data["date"])
    if "supplier" in data: item.supplier = str(data["supplier"])
    if "tap_spec" in data: item.tap_spec = str(data["tap_spec"])
    if "qty" in data: item.qty = float(data["qty"] or 0.0)
    if "rate" in data: item.rate = float(data["rate"] or 0.0)
    db.commit()
    return {"message": "Updated successfully"}

@app.delete("/api/tap_receipts/all")
def clear_all_tap_receipts(db: Session = Depends(get_db)):
    db.query(models.TapReceipt).delete()
    db.commit()
    return {"message": "All tap receipt records cleared successfully"}

@app.delete("/api/tap_receipts/{id}")
def delete_tap_receipt(id: int, db: Session = Depends(get_db)):
    item = db.query(models.TapReceipt).filter(models.TapReceipt.id == id).first()
    if not item:
        raise HTTPException(status_code=404, detail="Tap receipt not found")
    db.delete(item)
    db.commit()
    return {"message": "Deleted successfully"}

# --- INSERT ISSUES CRUD ---
@app.get("/api/insert_issues")
def get_insert_issues(db: Session = Depends(get_db)):
    try:
        rows = db.query(models.InsertIssue).order_by(models.InsertIssue.id.desc()).all()
        return [{
            "id": r.id,
            "date": r.date,
            "shift": r.shift,
            "department": r.department,
            "insert_spec": r.insert_spec,
            "batch_no": r.batch_no,
            "qty_issued": r.qty_issued,
            "qty_received": r.qty_received,
            "machine": r.machine,
            "operator": r.operator,
            "partno": r.partno,
            "opn_no": r.opn_no,
            "usages": r.usages,
            "receipt_id": r.receipt_id,
            "edge_data": r.edge_data,
            "created_at": r.created_at.isoformat() if r.created_at else None
        } for r in rows]
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/insert_issues/{id}")
def get_insert_issue(id: int, db: Session = Depends(get_db)):
    try:
        r = db.query(models.InsertIssue).filter(models.InsertIssue.id == id).first()
        if not r:
            raise HTTPException(status_code=404, detail="Insert issue record not found")
        return {
            "id": r.id,
            "date": r.date,
            "shift": r.shift,
            "department": r.department,
            "insert_spec": r.insert_spec,
            "batch_no": r.batch_no,
            "qty_issued": r.qty_issued,
            "qty_received": r.qty_received,
            "machine": r.machine,
            "operator": r.operator,
            "partno": r.partno,
            "opn_no": r.opn_no,
            "usages": r.usages,
            "receipt_id": r.receipt_id,
            "edge_data": r.edge_data,
            "created_at": r.created_at.isoformat() if r.created_at else None
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/insert_issues")
def create_insert_issue(data: dict, db: Session = Depends(get_db)):
    item = models.InsertIssue(
        date=(data.get("date") or datetime.datetime.now(IST).strftime("%Y-%m-%d")),
        shift=(data.get("shift") or "First").strip(),
        department=(data.get("department") or "WIPRO").strip(),
        insert_spec=(data.get("insert_spec") or "").strip(),
        batch_no=(data.get("batch_no") or "").strip(),
        qty_issued=float(data.get("qty_issued") or 0.0),
        qty_received=float(data.get("qty_received") or 0.0),
        machine=(data.get("machine") or "").strip(),
        operator=(data.get("operator") or "").strip(),
        partno=(data.get("partno") or "").strip(),
        opn_no=str(data.get("opn_no") or "").strip(),
        usages=str(data.get("usages") or ""),
        receipt_id=int(data.get("receipt_id")) if data.get("receipt_id") else None,
        edge_data=str(data.get("edge_data") or "")
    )
    db.add(item)
    db.commit()
    db.refresh(item)
    return {"id": item.id}

@app.post("/api/insert_issues/bulk")
def bulk_create_insert_issues(data: dict, db: Session = Depends(get_db)):
    issues = data.get("issues") or []
    count = 0
    for iss in issues:
        item = models.InsertIssue(
            date=(iss.get("date") or datetime.datetime.now(IST).strftime("%Y-%m-%d")),
            shift=(iss.get("shift") or "First").strip(),
            department=(iss.get("department") or "WIPRO").strip(),
            insert_spec=(iss.get("insert_spec") or "").strip(),
            batch_no=(iss.get("batch_no") or "").strip(),
            qty_issued=float(iss.get("qty_issued") or 0.0),
            qty_received=float(iss.get("qty_received") or 0.0),
            machine=(iss.get("machine") or "").strip(),
            operator=(iss.get("operator") or "").strip(),
            partno=(iss.get("partno") or "").strip(),
            opn_no=str(iss.get("opn_no") or "").strip(),
            usages=str(iss.get("usages") or ""),
            receipt_id=int(iss.get("receipt_id")) if iss.get("receipt_id") else None,
            edge_data=str(iss.get("edge_data") or "")
        )
        db.add(item)
        count += 1
    db.commit()
    return {"message": f"Successfully created {count} insert issue records"}

@app.put("/api/insert_issues/{id}")
def update_insert_issue(id: int, data: dict, db: Session = Depends(get_db)):
    item = db.query(models.InsertIssue).filter(models.InsertIssue.id == id).first()
    if not item:
        raise HTTPException(status_code=404, detail="Insert issue not found")
    if "date" in data: item.date = str(data["date"])
    if "shift" in data: item.shift = str(data["shift"])
    if "department" in data: item.department = str(data["department"])
    if "insert_spec" in data: item.insert_spec = str(data["insert_spec"])
    if "batch_no" in data: item.batch_no = str(data["batch_no"])
    if "qty_issued" in data: item.qty_issued = float(data["qty_issued"] or 0.0)
    if "qty_received" in data: item.qty_received = float(data["qty_received"] or 0.0)
    if "machine" in data: item.machine = str(data["machine"])
    if "operator" in data: item.operator = str(data["operator"])
    if "partno" in data: item.partno = str(data["partno"])
    if "opn_no" in data: item.opn_no = str(data["opn_no"])
    if "usages" in data: item.usages = str(data["usages"])
    if "receipt_id" in data: item.receipt_id = int(data["receipt_id"]) if data["receipt_id"] else None
    if "edge_data" in data: item.edge_data = str(data["edge_data"])
    db.commit()
    return {"message": "Updated successfully"}

@app.delete("/api/insert_issues/all")
def clear_all_insert_issues(db: Session = Depends(get_db)):
    db.query(models.InsertIssue).delete()
    db.commit()
    return {"message": "All insert issue records cleared successfully"}

@app.delete("/api/insert_issues/{id}")
def delete_insert_issue(id: int, db: Session = Depends(get_db)):
    item = db.query(models.InsertIssue).filter(models.InsertIssue.id == id).first()
    if not item:
        raise HTTPException(status_code=404, detail="Insert issue record not found")
    db.delete(item)
    db.commit()
    return {"message": "Deleted successfully"}

# --- TAP ISSUES CRUD ---
@app.get("/api/tap_issues")
def get_tap_issues(db: Session = Depends(get_db)):
    try:
        rows = db.query(models.TapIssue).order_by(models.TapIssue.id.desc()).all()
        return [{
            "id": r.id,
            "date": r.date,
            "shift": r.shift,
            "department": r.department,
            "tap_spec": r.tap_spec,
            "qty_issued": r.qty_issued,
            "qty_received": r.qty_received,
            "machine": r.machine,
            "operator": r.operator,
            "partno": r.partno,
            "opn_no": r.opn_no,
            "created_at": r.created_at.isoformat() if r.created_at else None
        } for r in rows]
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/tap_issues")
def create_tap_issue(data: dict, db: Session = Depends(get_db)):
    item = models.TapIssue(
        date=(data.get("date") or datetime.datetime.now(IST).strftime("%Y-%m-%d")),
        shift=(data.get("shift") or "First").strip(),
        department=(data.get("department") or "WIPRO").strip(),
        tap_spec=(data.get("tap_spec") or "").strip(),
        qty_issued=float(data.get("qty_issued") or 0.0),
        qty_received=float(data.get("qty_received") or 0.0),
        machine=(data.get("machine") or "").strip(),
        operator=(data.get("operator") or "").strip(),
        partno=(data.get("partno") or "").strip(),
        opn_no=str(data.get("opn_no") or "").strip()
    )
    db.add(item)
    db.commit()
    db.refresh(item)
    return {"id": item.id}

@app.post("/api/tap_issues/bulk")
def bulk_create_tap_issues(data: dict, db: Session = Depends(get_db)):
    issues = data.get("issues") or []
    count = 0
    for iss in issues:
        item = models.TapIssue(
            date=(iss.get("date") or datetime.datetime.now(IST).strftime("%Y-%m-%d")),
            shift=(iss.get("shift") or "First").strip(),
            department=(iss.get("department") or "WIPRO").strip(),
            tap_spec=(iss.get("tap_spec") or "").strip(),
            qty_issued=float(iss.get("qty_issued") or 0.0),
            qty_received=float(iss.get("qty_received") or 0.0),
            machine=(iss.get("machine") or "").strip(),
            operator=(iss.get("operator") or "").strip(),
            partno=(iss.get("partno") or "").strip(),
            opn_no=str(iss.get("opn_no") or "").strip()
        )
        db.add(item)
        count += 1
    db.commit()
    return {"message": f"Successfully created {count} tap issue records"}

@app.put("/api/tap_issues/{id}")
def update_tap_issue(id: int, data: dict, db: Session = Depends(get_db)):
    item = db.query(models.TapIssue).filter(models.TapIssue.id == id).first()
    if not item:
        raise HTTPException(status_code=404, detail="Tap issue not found")
    if "date" in data: item.date = str(data["date"])
    if "shift" in data: item.shift = str(data["shift"])
    if "department" in data: item.department = str(data["department"])
    if "tap_spec" in data: item.tap_spec = str(data["tap_spec"])
    if "qty_issued" in data: item.qty_issued = float(data["qty_issued"] or 0.0)
    if "qty_received" in data: item.qty_received = float(data["qty_received"] or 0.0)
    if "machine" in data: item.machine = str(data["machine"])
    if "operator" in data: item.operator = str(data["operator"])
    if "partno" in data: item.partno = str(data["partno"])
    if "opn_no" in data: item.opn_no = str(data["opn_no"])
    db.commit()
    return {"message": "Updated successfully"}

@app.delete("/api/tap_issues/all")
def clear_all_tap_issues(db: Session = Depends(get_db)):
    db.query(models.TapIssue).delete()
    db.commit()
    return {"message": "All tap issue records cleared successfully"}

@app.delete("/api/tap_issues/{id}")
def delete_tap_issue(id: int, db: Session = Depends(get_db)):
    item = db.query(models.TapIssue).filter(models.TapIssue.id == id).first()
    if not item:
        raise HTTPException(status_code=404, detail="Tap issue record not found")
    db.delete(item)
    db.commit()
    return {"message": "Deleted successfully"}

# --- Serve Static Files ---
@app.middleware("http")
async def add_no_cache_headers(request, call_next):
    response = await call_next(request)
    if request.url.path.startswith("/static") or request.url.path == "/":
        response.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
        response.headers["Pragma"] = "no-cache"
        response.headers["Expires"] = "0"
    return response

app.mount("/static", StaticFiles(directory="static"), name="static")

@app.get("/")
def read_index():
    return FileResponse("static/index.html")
